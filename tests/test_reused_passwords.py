"""REUSED PASSWORDS (Reports) — group all accounts / DPAPI entries sharing one secret."""

import collector.db as db


def _ws(name):
    with db.db_cursor() as cur:
        cur.execute("INSERT INTO workspaces(name) VALUES(?)", (name,))
        return cur.execute("SELECT id FROM workspaces WHERE name=?", (name,)).fetchone()["id"]


def _cred(cur, wid, domain, user, pw, credtype="plaintext", brut=None):
    cur.execute("INSERT INTO credentials(workspace_id,proto,domain,username,password,credtype,brutforced)"
                " VALUES(?,?,?,?,?,?,?)", (wid, "SMB", domain, user, pw, credtype, brut))


def _custom(cur, wid, domain, login, pw, credtype="plaintext", url=None):
    cur.execute("INSERT INTO custom_credentials(workspace_id,domain,login,password,credtype,url)"
                " VALUES(?,?,?,?,?,?)", (wid, domain, login, pw, credtype, url))


def _dpapi(cur, wid, user, pw, url=None):
    cur.execute("INSERT INTO dpapi_secrets(workspace_id,username,password,url) VALUES(?,?,?,?)",
                (wid, user, pw, url))


def _find(wid, min_count=2):
    from collector.services.reused_password_service import find_reused_passwords
    with db.db_cursor() as cur:
        return find_reused_passwords(cur, wid, min_count=min_count)


def test_reused_plaintext(auth_client):
    wid = _ws("rp-plain")
    with db.db_cursor() as cur:
        _cred(cur, wid, "CORP", "jdoe", "P@ss1")
        _cred(cur, wid, "WS01", "admin", "P@ss1")
        _cred(cur, wid, "CORP", "alice", "Unique1")   # not reused
    rows = _find(wid)
    assert len(rows) == 1
    r = rows[0]
    assert r["secret"] == "P@ss1" and r["type"] == "plaintext" and r["count"] == 2
    assert set(r["accounts"]) == {"CORP\\jdoe", "WS01\\admin"}


def test_cracked_hash_unifies_with_plaintext(auth_client):
    wid = _ws("rp-crack")
    with db.db_cursor() as cur:
        _cred(cur, wid, "CORP", "jdoe", "P@ss1", "plaintext")
        _cred(cur, wid, "CORP", "svc", "aad3b:NTHASH", "hash", brut="P@ss1")
    rows = _find(wid)
    assert len(rows) == 1 and rows[0]["secret"] == "P@ss1" and rows[0]["type"] == "plaintext"
    assert set(rows[0]["accounts"]) == {"CORP\\jdoe", "CORP\\svc"}


def test_uncracked_hash_reuse(auth_client):
    wid = _ws("rp-uhash")
    with db.db_cursor() as cur:
        _cred(cur, wid, "WS01", "Administrator", "aad3b:SHARED", "hash")
        _cred(cur, wid, "WS02", "Administrator", "aad3b:SHARED", "hash")
    rows = _find(wid)
    assert len(rows) == 1 and rows[0]["type"] == "hash" and rows[0]["count"] == 2


def test_dpapi_same_password(auth_client):
    wid = _ws("rp-dpapi")
    with db.db_cursor() as cur:
        _cred(cur, wid, "CORP", "jdoe", "P@ss1")
        _dpapi(cur, wid, "jdoe", "P@ss1", url="https://site")
    rows = _find(wid)
    assert len(rows) == 1
    r = rows[0]
    assert r["count"] == 2
    assert r["accounts"] == ["CORP\\jdoe"]
    assert r["dpapi"] == ["https://site;jdoe"]


def test_empty_password_excluded(auth_client):
    wid = _ws("rp-empty")
    with db.db_cursor() as cur:
        _cred(cur, wid, "CORP", "a", "<empty_password>")
        _cred(cur, wid, "CORP", "b", "<empty_password>")
    assert _find(wid) == []


def test_whitespace_only_password_excluded(auth_client):
    # spaces/tabs are "blank" too — must not create a row (dpapi passwords aren't normalized)
    wid = _ws("rp-ws")
    with db.db_cursor() as cur:
        _dpapi(cur, wid, "", "   ", url="https://a")
        _dpapi(cur, wid, "", "   ", url="https://b")
        _cred(cur, wid, "CORP", "x", "  ")
        _cred(cur, wid, "CORP", "y", "  ")
    assert _find(wid) == []


def test_guest_excluded(auth_client):
    wid = _ws("rp-guest")
    with db.db_cursor() as cur:
        _cred(cur, wid, "CORP", "Guest", "P@ss1")
        _cred(cur, wid, "CORP", "jdoe", "P@ss1")   # only jdoe → count 1 → excluded
    assert _find(wid) == []


def test_custom_without_login(auth_client):
    wid = _ws("rp-custom")
    with db.db_cursor() as cur:
        _cred(cur, wid, "CORP", "jdoe", "Shared9")
        _custom(cur, wid, "", "", "Shared9", url="http://app")   # custom: no login/domain
    rows = _find(wid)
    assert len(rows) == 1 and rows[0]["count"] == 2
    assert "(no login)" in rows[0]["accounts"]


def test_reused_passwords_api_and_export(auth_client):
    auth_client.post("/api/workspaces", json={"name": "rp-api"})
    with db.db_cursor() as cur:
        wid = cur.execute("SELECT id FROM workspaces WHERE name='rp-api'").fetchone()["id"]
        _cred(cur, wid, "CORP", "jdoe", "P@ss1")
        _cred(cur, wid, "WS01", "admin", "P@ss1")
    r = auth_client.get(f"/api/reports/reused-passwords?workspace_id={wid}")
    assert r.status_code == 200 and r.json()["count"] == 1
    e = auth_client.get(f"/api/reports/reused-passwords/export?workspace_id={wid}")
    assert e.status_code == 200
    assert "rp-api_reused_passwords.xlsx" in e.headers["content-disposition"]
    assert e.content[:2] == b"PK"
    import io
    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(e.content)).active
    hdr = [c.value for c in ws[1]]
    assert hdr == ["Password / Hash", "Type", "Accounts (domain\\login)", "DPAPI (url;login)", "Count"]
    accounts_cell = ws.cell(row=2, column=3)          # Accounts column
    assert "CORP\\jdoe" in accounts_cell.value and "WS01\\admin" in accounts_cell.value
    assert "\n" in accounts_cell.value                 # one entry per line
    assert accounts_cell.alignment.wrap_text is True   # so Excel renders the lines


def test_reused_button_in_reports(auth_client):
    html = auth_client.get("/api/shell/module/reports/ui").text
    assert "REUSED PASSWORDS" in html
    assert "RPModule.exportReusedPasswords()" in html
    assert html.index("DOWNLOAD TIMELINE") < html.index("REUSED PASSWORDS")
