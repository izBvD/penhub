"""LOCAL ADMIN FOUNDER — detection service + API."""

import collector.db as db


def _mk_ws(name):
    with db.db_cursor() as cur:
        cur.execute("INSERT INTO workspaces(name) VALUES(?)", (name,))
        return cur.execute("SELECT id FROM workspaces WHERE name=?", (name,)).fetchone()["id"]


def _host(cur, wid, ip, hostname, domain, dc=0):
    cur.execute("INSERT INTO hosts(workspace_id,ip,hostname,domain,dc) VALUES(?,?,?,?,?)",
                (wid, ip, hostname, domain, dc))
    return cur.execute("SELECT id FROM hosts WHERE workspace_id=? AND ip=?", (wid, ip)).fetchone()["id"]


def _cred(cur, wid, domain, username, password, credtype="hash", pfip=None,
          local_admin_cred=0, brutforced=None):
    cur.execute(
        "INSERT INTO credentials(workspace_id,proto,domain,username,password,credtype,"
        "pillaged_from_ip,local_admin_cred,brutforced) VALUES(?,?,?,?,?,?,?,?,?)",
        (wid, "SMB", domain, username, password, credtype, pfip, local_admin_cred, brutforced))
    return cur.execute("SELECT id FROM credentials WHERE workspace_id=? AND domain=? AND username=? "
                       "AND password=?", (wid, domain, username, password)).fetchone()["id"]


def _admin_rel(cur, wid, cred_id, host_id):
    cur.execute("INSERT INTO auth_relations(workspace_id,proto,credential_id,host_id,relation_type)"
                " VALUES(?,?,?,?,?)", (wid, "SMB", cred_id, host_id, "admin"))


def test_reuse_local_admin_detected(auth_client):
    wid = _mk_ws("la-reuse")
    with db.db_cursor() as cur:
        _host(cur, wid, "10.0.0.1", "WS01", "CORP", dc=0)
        _host(cur, wid, "10.0.0.2", "WS02", "CORP", dc=0)
        _host(cur, wid, "10.0.0.9", "DC01", "CORP", dc=1)   # makes CORP an AD domain
        # same local Administrator (hostname domain) on 2 machines, same hash
        _cred(cur, wid, "WS01", "Administrator", "aad3b:31d6HASH", pfip="10.0.0.1")
        _cred(cur, wid, "WS02", "Administrator", "aad3b:31d6HASH", pfip="10.0.0.2")
    from collector.services.local_admin_service import find_local_admins
    with db.db_cursor() as cur:
        rows = find_local_admins(cur, wid, min_hosts=2)
    assert len(rows) == 1
    r = rows[0]
    assert r["username"] == "Administrator" and r["machine_count"] == 2
    assert set(r["machines"]) == {"WS01", "WS02"} and r["tier"] == "reuse"


def test_domain_account_excluded(auth_client):
    wid = _mk_ws("la-domain")
    with db.db_cursor() as cur:
        _host(cur, wid, "10.0.1.1", "WS01", "CORP")
        _host(cur, wid, "10.0.1.2", "WS02", "CORP")
        _host(cur, wid, "10.0.1.9", "DC01", "CORP", dc=1)
        # domain account: domain=CORP (AD domain). Same (domain,user,hash) from many hosts
        # collapses to ONE row (UNIQUE key) — that is the real behaviour. Must be excluded.
        _cred(cur, wid, "CORP", "jdoe", "aad3b:DOMHASH", pfip="10.0.1.1")
    from collector.services.local_admin_service import find_local_admins
    with db.db_cursor() as cur:
        rows = find_local_admins(cur, wid, min_hosts=2)
    assert rows == []


def test_collision_local_kept_domain_dropped(auth_client):
    wid = _mk_ws("la-collide")
    with db.db_cursor() as cur:
        _host(cur, wid, "10.0.2.1", "WS01", "CORP")
        _host(cur, wid, "10.0.2.2", "WS02", "CORP")
        _host(cur, wid, "10.0.2.9", "DC01", "CORP", dc=1)
        _cred(cur, wid, "CORP", "Administrator", "aad3b:DOMAINPW", pfip="10.0.2.9")  # domain admin
        _cred(cur, wid, "WS01", "Administrator", "aad3b:LOCALPW", pfip="10.0.2.1")   # local
        _cred(cur, wid, "WS02", "Administrator", "aad3b:LOCALPW", pfip="10.0.2.2")   # local (reuse)
    from collector.services.local_admin_service import find_local_admins
    with db.db_cursor() as cur:
        rows = find_local_admins(cur, wid, min_hosts=2)
    assert len(rows) == 1
    assert rows[0]["secret"] == "aad3b:LOCALPW" and rows[0]["machine_count"] == 2


def test_reuse_detected_without_pillaged_ip(auth_client):
    # Regression: LSA+SAM creds often have pillaged_from_ip=NULL and no auth_relation
    # (nxc didn't tag the source host). They are visible in the LSA+SAM tab and MUST
    # still be detected as reused local creds — machine count comes from domain, not pfip.
    wid = _mk_ws("la-nopf")
    with db.db_cursor() as cur:
        _host(cur, wid, "10.7.0.9", "DC01", "CORP", dc=1)   # makes CORP an AD domain
        _cred(cur, wid, "WS01", "Administrator", "aad3b:NOPF", pfip=None)
        _cred(cur, wid, "WS02", "Administrator", "aad3b:NOPF", pfip=None)
    from collector.services.local_admin_service import find_local_admins
    with db.db_cursor() as cur:
        rows = find_local_admins(cur, wid, min_hosts=2)
    assert len(rows) == 1 and rows[0]["tier"] == "reuse" and rows[0]["machine_count"] == 2


def test_operator_marked_always_included(auth_client):
    wid = _mk_ws("la-oper")
    with db.db_cursor() as cur:
        _host(cur, wid, "10.0.3.1", "WS01", "CORP")
        _host(cur, wid, "10.0.3.9", "DC01", "CORP", dc=1)
        _cred(cur, wid, "WS01", "svc_local", "aad3b:SVC", pfip="10.0.3.1", local_admin_cred=1)
    from collector.services.local_admin_service import find_local_admins
    with db.db_cursor() as cur:
        rows = find_local_admins(cur, wid, min_hosts=2)
    assert len(rows) == 1 and rows[0]["tier"] == "operator"


def test_admin_relation_single_host_included(auth_client):
    wid = _mk_ws("la-admin")
    with db.db_cursor() as cur:
        h = _host(cur, wid, "10.0.4.1", "WS01", "CORP")
        _host(cur, wid, "10.0.4.9", "DC01", "CORP", dc=1)
        cid = _cred(cur, wid, "WS01", "localadm", "aad3b:LA", pfip="10.0.4.1")
        _admin_rel(cur, wid, cid, h)   # proven admin on SMB
    from collector.services.local_admin_service import find_local_admins
    with db.db_cursor() as cur:
        rows = find_local_admins(cur, wid, min_hosts=2)
    assert len(rows) == 1 and rows[0]["tier"] == "admin" and rows[0]["machine_count"] == 1


def test_single_local_no_admin_not_reused_excluded(auth_client):
    wid = _mk_ws("la-single")
    with db.db_cursor() as cur:
        _host(cur, wid, "10.0.5.1", "WS01", "CORP")
        _host(cur, wid, "10.0.5.9", "DC01", "CORP", dc=1)
        _cred(cur, wid, "WS01", "randuser", "aad3b:RU", pfip="10.0.5.1")  # 1 machine, no admin
    from collector.services.local_admin_service import find_local_admins
    with db.db_cursor() as cur:
        rows = find_local_admins(cur, wid, min_hosts=2)
    assert rows == []


def test_local_admins_api_json_and_export(auth_client):
    wid = _mk_ws("la-api")
    with db.db_cursor() as cur:
        _host(cur, wid, "10.9.0.1", "WS01", "CORP")
        _host(cur, wid, "10.9.0.2", "WS02", "CORP")
        _host(cur, wid, "10.9.0.9", "DC01", "CORP", dc=1)
        _cred(cur, wid, "WS01", "Administrator", "aad3b:REUSE", pfip="10.9.0.1")
        _cred(cur, wid, "WS02", "Administrator", "aad3b:REUSE", pfip="10.9.0.2")
    r = auth_client.get(f"/api/reports/local-admins?workspace_id={wid}")
    assert r.status_code == 200
    j = r.json()
    assert j["count"] == 1 and j["rows"][0]["username"] == "Administrator"
    e = auth_client.get(f"/api/reports/local-admins/export?workspace_id={wid}")
    assert e.status_code == 200
    assert "la-api_local_admins.xlsx" in e.headers["content-disposition"]
    assert e.content[:2] == b"PK"   # xlsx = zip


def test_export_two_sections(auth_client):
    import io
    import openpyxl
    wid = _mk_ws("la-sections")
    with db.db_cursor() as cur:
        h = _host(cur, wid, "10.8.0.1", "WS01", "CORP")
        _host(cur, wid, "10.8.0.2", "WS02", "CORP")
        _host(cur, wid, "10.8.0.9", "DC01", "CORP", dc=1)
        # admin-proven local admin (tier=admin)
        cid = _cred(cur, wid, "WS01", "adminacct", "aad3b:ADMIN", pfip="10.8.0.1")
        _admin_rel(cur, wid, cid, h)
        # reused non-admin local cred on 2 machines (tier=reuse)
        _cred(cur, wid, "WS01", "svcuser", "aad3b:SHARED", pfip="10.8.0.1")
        _cred(cur, wid, "WS02", "svcuser", "aad3b:SHARED", pfip="10.8.0.2")
    e = auth_client.get(f"/api/reports/local-admins/export?workspace_id={wid}")
    wb = openpyxl.load_workbook(io.BytesIO(e.content))
    colA = [(r[0].value or "") for r in wb.active.iter_rows()]
    i_admins = colA.index("LOCAL ADMINS")
    i_reused = colA.index("REUSED LOCAL CREDENTIALS")
    assert i_admins < i_reused
    i_adminacct = colA.index("adminacct")
    i_svcuser = colA.index("svcuser")
    assert i_admins < i_adminacct < i_reused          # admin in first section
    assert i_reused < i_svcuser                        # reused in second section

    # Domain column dropped from BOTH sections (Machine list already lists the machines)
    all_rows = list(wb.active.iter_rows(values_only=True))
    admin_hdr = all_rows[i_admins + 1]
    reuse_hdr = all_rows[i_reused + 1]
    assert "Domain" not in admin_hdr and "Domain" not in reuse_hdr
    assert "Machine list" in admin_hdr and "Machine list" in reuse_hdr


def test_local_admin_block_between_timeline_and_exports(auth_client):
    html = auth_client.get("/api/shell/module/reports/ui").text
    assert "LOCAL ADMIN FOUNDER" in html
    assert "RPModule.exportLocalAdmins()" in html
    # placed between TIMELINE and EXPORTS blocks
    assert html.index("TIMELINE") < html.index("LOCAL ADMIN FOUNDER") < html.index("EXPORTS")


def test_guest_and_empty_excluded(auth_client):
    wid = _mk_ws("la-guest")
    with db.db_cursor() as cur:
        _host(cur, wid, "10.0.6.1", "WS01", "CORP")
        _host(cur, wid, "10.0.6.2", "WS02", "CORP")
        _host(cur, wid, "10.0.6.9", "DC01", "CORP", dc=1)
        _cred(cur, wid, "WS01", "Guest", "aad3b:G", pfip="10.0.6.1")
        _cred(cur, wid, "WS02", "Guest", "aad3b:G", pfip="10.0.6.2")
        _cred(cur, wid, "WS01", "Administrator", "<empty_password>", credtype="plaintext", pfip="10.0.6.1")
        _cred(cur, wid, "WS02", "Administrator", "<empty_password>", credtype="plaintext", pfip="10.0.6.2")
    from collector.services.local_admin_service import find_local_admins
    with db.db_cursor() as cur:
        rows = find_local_admins(cur, wid, min_hosts=2)
    assert rows == []
