"""
Smoke tests — run after every change.
Auth, workspace CRUD, API endpoint availability.
No sync payload needed — all endpoints checked with an empty workspace.
"""

import io

from fastapi.testclient import TestClient

from penhub.app import app
from tests.conftest import TEST_PASSWORD


def test_login_ok():
    with TestClient(app) as c:
        r = c.post("/api/login", json={"password": TEST_PASSWORD})
    assert r.status_code == 200
    assert r.json()["ok"] is True


def test_login_wrong():
    with TestClient(app) as c:
        r = c.post("/api/login", json={"password": "wrong"})
    assert r.status_code == 403


def test_workspaces_unauthenticated():
    with TestClient(app) as c:
        r = c.get("/api/workspaces")
    assert r.status_code == 401


def test_root_page(auth_client):
    r = auth_client.get("/")
    assert r.status_code == 200
    assert "PenHub" in r.text


def test_workspace_create_and_list(auth_client, workspace_id):
    ws_list = auth_client.get("/api/workspaces").json()
    assert any(w["name"] == "test-ws" for w in ws_list)


def test_workspace_create_duplicate_returns_409(auth_client):
    import uuid
    name = f"__dup_test_{uuid.uuid4().hex[:8]}__"
    r1 = auth_client.post("/api/workspaces", json={"name": name})
    assert r1.status_code == 200, "first creation must succeed"

    r2 = auth_client.post("/api/workspaces", json={"name": name})
    assert r2.status_code == 409
    assert "already exists" in r2.json().get("detail", "").lower()

    from collector.db import db_cursor
    with db_cursor() as cur:
        cur.execute("DELETE FROM workspaces WHERE name=?", (name,))


def test_workspace_create_unique_name_succeeds(auth_client):
    import uuid
    unique = f"__test_{uuid.uuid4().hex[:8]}__"
    r = auth_client.post("/api/workspaces", json={"name": unique})
    assert r.status_code == 200
    data = r.json()
    assert data["name"] == unique
    assert "id" in data
    # Cleanup
    from collector.db import db_cursor
    with db_cursor() as cur:
        cur.execute("DELETE FROM workspaces WHERE name=?", (unique,))


def test_workspace_create_empty_name_rejected(auth_client):
    r = auth_client.post("/api/workspaces", json={"name": "   "})
    assert r.status_code == 400


class TestRecycleBin:
    """Soft-delete / recycle-bin workflow for workspaces."""

    def _create(self, auth_client, name):
        r = auth_client.post("/api/workspaces", json={"name": name})
        assert r.status_code == 200
        return r.json()["id"]

    def _hard_delete(self, ws_id):
        from collector.db import db_cursor
        with db_cursor() as cur:
            cur.execute("DELETE FROM workspaces WHERE id=?", (ws_id,))

    def test_delete_moves_to_recycle(self, auth_client):
        """DELETE /api/workspaces/{id} soft-deletes: workspace still exists, recycled_at set."""
        import uuid
        ws_id = self._create(auth_client, f"__rc_{uuid.uuid4().hex[:6]}__")
        try:
            r = auth_client.delete(f"/api/workspaces/{ws_id}")
            assert r.status_code == 200
            assert r.json()["ok"] is True

            from collector.db import db_cursor
            with db_cursor() as cur:
                row = cur.execute(
                    "SELECT recycled_at FROM workspaces WHERE id=?", (ws_id,)
                ).fetchone()
            assert row is not None, "workspace must still exist after soft-delete"
            assert row["recycled_at"] is not None, "recycled_at must be set"
        finally:
            self._hard_delete(ws_id)

    def test_recycled_appears_with_recycled_at_in_list(self, auth_client):
        """GET /api/workspaces returns recycled workspace with recycled_at field set."""
        import uuid
        ws_id = self._create(auth_client, f"__rc_{uuid.uuid4().hex[:6]}__")
        try:
            auth_client.delete(f"/api/workspaces/{ws_id}")
            ws_list = auth_client.get("/api/workspaces").json()
            recycled = next((w for w in ws_list if w["id"] == ws_id), None)
            assert recycled is not None, "recycled workspace must appear in list"
            assert recycled["recycled_at"] is not None
        finally:
            self._hard_delete(ws_id)

    def test_recycle_name_still_unique(self, auth_client):
        """Recycled workspace still holds its name — creating duplicate returns 409."""
        import uuid
        name = f"__rc_{uuid.uuid4().hex[:6]}__"
        ws_id = self._create(auth_client, name)
        try:
            auth_client.delete(f"/api/workspaces/{ws_id}")
            r = auth_client.post("/api/workspaces", json={"name": name})
            assert r.status_code == 409
        finally:
            self._hard_delete(ws_id)

    def test_restore_to_active(self, auth_client):
        """POST /restore_active: clears both recycled_at and archived_at."""
        import uuid
        ws_id = self._create(auth_client, f"__rc_{uuid.uuid4().hex[:6]}__")
        try:
            auth_client.post(f"/api/workspaces/{ws_id}/archive")
            auth_client.delete(f"/api/workspaces/{ws_id}")

            r = auth_client.post(f"/api/workspaces/{ws_id}/restore_active")
            assert r.status_code == 200

            from collector.db import db_cursor
            with db_cursor() as cur:
                row = cur.execute(
                    "SELECT recycled_at, archived_at FROM workspaces WHERE id=?", (ws_id,)
                ).fetchone()
            assert row["recycled_at"] is None
            assert row["archived_at"] is None
        finally:
            self._hard_delete(ws_id)

    def test_restore_to_archive(self, auth_client):
        """POST /restore_archive: clears recycled_at, keeps/sets archived_at."""
        import uuid
        ws_id = self._create(auth_client, f"__rc_{uuid.uuid4().hex[:6]}__")
        try:
            auth_client.delete(f"/api/workspaces/{ws_id}")

            r = auth_client.post(f"/api/workspaces/{ws_id}/restore_archive")
            assert r.status_code == 200

            from collector.db import db_cursor
            with db_cursor() as cur:
                row = cur.execute(
                    "SELECT recycled_at, archived_at FROM workspaces WHERE id=?", (ws_id,)
                ).fetchone()
            assert row["recycled_at"] is None
            assert row["archived_at"] is not None
        finally:
            self._hard_delete(ws_id)

    def test_permanent_delete(self, auth_client):
        """DELETE /permanent: hard-deletes workspace, data is gone."""
        import uuid
        ws_id = self._create(auth_client, f"__rc_{uuid.uuid4().hex[:6]}__")
        auth_client.delete(f"/api/workspaces/{ws_id}")

        r = auth_client.delete(f"/api/workspaces/{ws_id}/permanent")
        assert r.status_code == 200

        from collector.db import db_cursor
        with db_cursor() as cur:
            row = cur.execute("SELECT id FROM workspaces WHERE id=?", (ws_id,)).fetchone()
        assert row is None, "workspace must be gone after permanent delete"

    def test_permanent_delete_enrich_before_delete(self, auth_client):
        """Smart enrich must run BEFORE CASCADE delete so plaintext pairs are saved to HK."""
        import uuid
        import collector.hashkiller_db as hk_db
        from collector.nt_hash import nt_hash

        name  = f"__rc_r162_{uuid.uuid4().hex[:6]}__"
        plain = f"R162pass_{uuid.uuid4().hex[:4]}!"
        ws_id = self._create(auth_client, name)

        # Sync a plaintext credential into the workspace
        auth_client.post("/api/sync", json={
            "workspace": name, "operator": "tester",
            "data": {
                "hosts": [], "auth_relations": [], "dpapi_secrets": [],
                "shares": [], "ssh_keys": [], "conf_checks_results": [],
                "directory_listings": [], "vuln_findings": [],
                "credentials": [{"proto": "SMB", "domain": "", "username": "u_r162",
                                  "password": plain, "credtype": "plaintext"}],
            },
        })

        # Move to recycle, then permanently delete
        auth_client.delete(f"/api/workspaces/{ws_id}")
        r = auth_client.delete(f"/api/workspaces/{ws_id}/permanent")
        assert r.status_code == 200

        # Workspace must be gone
        from collector.db import db_cursor
        with db_cursor() as cur:
            row = cur.execute("SELECT id FROM workspaces WHERE id=?", (ws_id,)).fetchone()
        assert row is None

        # NT pair must be in HK — proves enrich ran before credentials were deleted
        nh = nt_hash(plain)
        pairs = hk_db.bulk_lookup({nh})
        assert nh in pairs, "smart enrich must save NT pair before permanent deletion"

        hk_db.delete_by_value(nh)


def test_stats_shape(auth_client, workspace_id):
    r = auth_client.get(f"/api/stats?workspace_id={workspace_id}")
    assert r.status_code == 200
    assert all(k in r.json() for k in ("hosts", "creds", "admin", "dpapi"))


def test_hosts_list_shape(auth_client, workspace_id):
    r = auth_client.get(f"/api/hosts?workspace_id={workspace_id}")
    assert r.status_code == 200
    d = r.json()
    assert "rows" in d and "total" in d


def test_credentials_list_shape(auth_client, workspace_id):
    r = auth_client.get(f"/api/credentials?workspace_id={workspace_id}")
    assert r.status_code == 200
    assert "rows" in r.json()


def test_dpapi_list_shape(auth_client, workspace_id):
    r = auth_client.get(f"/api/dpapi?workspace_id={workspace_id}")
    assert r.status_code == 200
    assert "rows" in r.json()


def test_hk_stats_shape(auth_client):
    # Fast stats: total + smart only (warning moved to /api/hk/stats/warning, computed lazily).
    r = auth_client.get("/api/hk/stats")
    assert r.status_code == 200
    assert all(k in r.json() for k in ("total", "smart"))

    rw = auth_client.get("/api/hk/stats/warning")
    assert rw.status_code == 200
    assert "warning" in rw.json()


def test_export_xlsx_all_view_no_query_binding_error(auth_client, workspace_id):
    """Regression: get_credentials called directly with missing params used Query(None)
    as default — truthy check 'if credtype:' added Query object to SQL params → SQLite error."""
    r = auth_client.get(f"/api/export/xlsx?workspace_id={workspace_id}&view=all")
    assert r.status_code == 200


def test_get_functions_direct_call_without_optional_params(workspace_id):
    """Regression: data.py get_* functions must work when called directly from Python
    without specifying optional params. If params use = Query(None) instead of = None,
    callers get a truthy FieldInfo object → silent SQL corruption."""
    from collector.api.data import get_credentials, get_dpapi, get_results, get_vulns
    from collector.api.data_hosts import get_hosts
    # Each call omits all optional params — only workspace_id is required.
    # Any Query() default would cause sqlite3.ProgrammingError or AttributeError here.
    result = get_credentials(workspace_id=workspace_id)
    assert "rows" in result
    result = get_results(workspace_id=workspace_id)
    assert "rows" in result
    result = get_hosts(workspace_id=workspace_id)
    assert "rows" in result
    result = get_dpapi(workspace_id=workspace_id)
    assert "rows" in result
    result = get_vulns(workspace_id=workspace_id)
    assert "rows" in result


def test_custom_import_template_download(auth_client):
    """Smoke: template endpoint returns a valid xlsx file with correct columns."""
    import openpyxl, io
    r = auth_client.get("/api/toolbox/custom-import/template")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, 11)]
    assert headers == ["Proto", "IP", "Port", "Domain", "Login", "Password",
                       "Type", "URL", "Source", "Comment"]
    # Login (col 5) and Password (col 6) must use the orange required-fill
    # Check last 6 hex chars — openpyxl alpha prefix varies by version (FF vs 00)
    assert ws.cell(row=1, column=5).fill.fgColor.rgb.upper().endswith("C55A11")
    assert ws.cell(row=1, column=6).fill.fgColor.rgb.upper().endswith("C55A11")
    # Other headers must use the standard blue fill
    assert ws.cell(row=1, column=1).fill.fgColor.rgb.upper().endswith("2F75B6")


def test_custom_import_upload(auth_client, workspace_id):
    """Smoke + regression: upload, enrichment, re-import, Russian chars, counters."""
    import openpyxl
    from collector.db import db_cursor

    HEADERS = ["Proto", "IP", "Port", "Domain", "Login", "Password",
               "Type", "URL", "Source", "Comment"]

    def _make_xlsx(*data_rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)
        for r in data_rows:
            ws.append(list(r))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _upload(xlsx_bytes, fname="test.xlsx"):
        return auth_client.post(
            "/api/toolbox/custom-import/upload",
            data={"workspace_id": str(workspace_id)},
            files={"file": (fname, xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    # Clean up before test (session-scoped workspace may have leftovers)
    with db_cursor() as cur:
        cur.execute("DELETE FROM custom_credentials WHERE workspace_id=?", (workspace_id,))

    try:
        # ── Round 1: fresh import ──────────────────────────────────────────────
        xlsx1 = _make_xlsx(
            # Proto  IP          Port  Domain  Login       Password     Type  URL  Source  Comment
            ["SMB", "10.0.0.1", 445,  "CORP", "alice",    "P@ssw0rd",  "",   "",  "",     ""],
            ["",    "",         "",   "",     "иванов",   "пароль123", "",   "",  "",     ""],  # Russian
            ["",    "",         "",   "",     "",         "",          "",   "",  "",     ""],  # skipped
        )
        r1 = _upload(xlsx1)
        assert r1.status_code == 200
        d1 = r1.json()
        assert d1["added"]           == 2, d1
        assert d1["skipped"]         == 1, d1
        assert d1["enriched"]        == 0, d1
        assert d1["already_existed"] == 0, d1

        # ── Round 2: exact re-import → already_existed ─────────────────────────
        r2 = _upload(xlsx1)
        assert r2.status_code == 200
        d2 = r2.json()
        assert d2["already_existed"] == 2, d2
        assert d2["added"]           == 0, d2

        # ── Round 3: enrichment — add IP to Russian cred ───────────────────────
        xlsx3 = _make_xlsx(
            ["RDP", "192.168.1.5", 3389, "", "иванов", "пароль123", "", "", "phishing", ""],
        )
        r3 = _upload(xlsx3)
        assert r3.status_code == 200
        d3 = r3.json()
        assert d3["enriched"] == 1, d3
        assert d3["added"]    == 0, d3

        # Verify enrichment actually written to DB
        with db_cursor() as cur:
            row = cur.execute(
                "SELECT proto, ip, port, source FROM custom_credentials"
                " WHERE workspace_id=? AND login=? AND password=?",
                (workspace_id, "иванов", "пароль123"),
            ).fetchone()
        assert row is not None
        assert row["proto"] == "RDP"
        assert row["ip"]    == "192.168.1.5"
        assert row["port"]  == 3389
        assert row["source"] == "phishing"

        # ── Round 4: same login+pass with NEW ip → new credential ─────────────
        # After Round 3: иванов/пароль123 stored as (RDP, 192.168.1.5, 3389).
        # Incoming has same proto=RDP but different ip=10.10.10.10
        # → stored ip is set and differs from incoming → incompatible → INSERT.
        xlsx4 = _make_xlsx(
            ["RDP", "10.10.10.10", 3389, "", "иванов", "пароль123", "", "", "", ""],
        )
        r4 = _upload(xlsx4)
        assert r4.status_code == 200
        d4 = r4.json()
        assert d4["added"] == 1, d4   # new IP → new credential

        # ── Round 5: same login+pass, DIFFERENT proto → must INSERT, not enrich ─
        # Both existing иванов rows have proto=RDP.
        # Incoming has proto=SMB → stored proto is set and conflicts → no enrichable candidate.
        # Must create a new (SMB, 192.168.1.5) credential, not silently reuse an RDP row.
        xlsx5 = _make_xlsx(
            ["SMB", "192.168.1.5", 445, "CORP", "иванов", "пароль123", "", "", "", ""],
        )
        r5 = _upload(xlsx5)
        assert r5.status_code == 200
        d5 = r5.json()
        assert d5["added"] == 1, d5      # different proto → new credential
        assert d5["enriched"] == 0, d5   # must NOT enrich an existing RDP row

        with db_cursor() as cur:
            smb_row = cur.execute(
                "SELECT proto, ip FROM custom_credentials"
                " WHERE workspace_id=? AND login=? AND password=? AND proto=?",
                (workspace_id, "иванов", "пароль123", "SMB"),
            ).fetchone()
        assert smb_row is not None, "SMB credential must be inserted separately"
        assert smb_row["ip"] == "192.168.1.5"

        # ── Round 6: same (proto, ip, login, pass) with DIFFERENT url → new cred ─
        # url is part of the UNIQUE key: same network context but different URL =
        # different application credential (e.g., two web endpoints on the same host).
        xlsx6 = _make_xlsx(
            ["HTTPS", "10.0.0.2", 443, "", "alice", "P@ssw0rd", "", "http://app1.example.com", "", ""],
        )
        r6 = _upload(xlsx6)
        assert r6.status_code == 200
        d6 = r6.json()
        assert d6["added"] == 1, d6   # new url → new credential

        xlsx7 = _make_xlsx(
            ["HTTPS", "10.0.0.2", 443, "", "alice", "P@ssw0rd", "", "http://app2.example.com", "", ""],
        )
        r7 = _upload(xlsx7)
        assert r7.status_code == 200
        d7 = r7.json()
        assert d7["added"] == 1, d7      # different url → different credential
        assert d7["enriched"] == 0, d7   # must NOT enrich the app1 row

        # Re-import same → already_existed
        r7b = _upload(xlsx7)
        assert r7b.status_code == 200
        assert r7b.json()["already_existed"] == 1

        # Enrich NULL-url cred by providing a url
        xlsx8_base = _make_xlsx(
            ["HTTPS", "10.0.0.3", 443, "", "bob", "secret", "", "", "", ""],
        )
        r8a = _upload(xlsx8_base)
        assert r8a.status_code == 200
        assert r8a.json()["added"] == 1

        xlsx8_enrich = _make_xlsx(
            ["HTTPS", "10.0.0.3", 443, "", "bob", "secret", "", "http://portal.example.com", "", ""],
        )
        r8b = _upload(xlsx8_enrich)
        assert r8b.status_code == 200
        assert r8b.json()["enriched"] == 1   # NULL url → filled with incoming

    finally:
        with db_cursor() as cur:
            cur.execute("DELETE FROM custom_credentials WHERE workspace_id=?", (workspace_id,))


def test_hk_warning_recalculated_after_delete():
    """Regression: after deleting one of two conflicting plaintexts for a hash,
    warning count and export must reflect the current state — not a stale stored flag.
    The stored warning=1 flag is intentionally NOT cleared on delete; instead
    get_stats/get_warning_pairs/bulk_lookup derive warning status from actual row count."""
    import collector.hashkiller_db as hk_db

    NH = "aabbccdd" * 4  # 32 hex chars, valid NT hash placeholder
    P1 = "__test_plain_one__"
    P2 = "__test_plain_two__"

    hk_db.init_hk_db()

    # Clean up any leftover state from previous runs
    hk_db.delete_by_value(NH)

    try:
        # Import two plaintexts for same hash → conflict → both warning=1
        hk_db.bulk_import(f"{NH}:{P1}\n{NH}:{P2}")

        assert hk_db.get_warning_count() == 2, "both rows must appear as warning before delete"

        warning_pairs = hk_db.get_warning_pairs()
        hashes_in_export = {r["nt_hash"] for r in warning_pairs}
        assert NH in hashes_in_export, "hash must appear in warning export before delete"

        # Delete one conflicting plaintext — leaves exactly one plaintext for this hash
        hk_db.delete_by_value(P2)

        assert hk_db.get_warning_count() == 0, (
            "warning count must be 0 after conflict is resolved by deletion"
        )

        warning_pairs_after = hk_db.get_warning_pairs()
        hashes_after = {r["nt_hash"] for r in warning_pairs_after}
        assert NH not in hashes_after, (
            "resolved hash must not appear in warning export after deletion"
        )

        # bulk_lookup must now return the surviving plaintext
        result = hk_db.bulk_lookup({NH})
        assert result.get(NH) == P1, "surviving plaintext must be returned by bulk_lookup"

    finally:
        hk_db.delete_by_value(NH)


# ═══════════════════════════════════════════════════════════════════════════════
# R7.2 — per-row deletion of Domain Admin Watchlist entries
# ═══════════════════════════════════════════════════════════════════════════════

class TestDalEntryDelete:
    def _ws(self, auth_client, name):
        return auth_client.post("/api/workspaces", json={"name": name}).json()["id"]

    def _upload(self, auth_client, ws_id, domain, *usernames):
        return auth_client.post("/api/domain_admin_list/upload",
                                json={"workspace_id": ws_id, "domain": domain,
                                      "usernames": list(usernames)})

    def _pending(self, auth_client, ws_id):
        r = auth_client.get(f"/api/domain_admin_list/pending?workspace_id={ws_id}")
        return {(row["domain"], row["username"]) for row in r.json()["rows"]}

    def _delete(self, auth_client, ws_id, domain, username):
        p = f"workspace_id={ws_id}&domain={domain}&username={username}"
        return auth_client.delete(f"/api/domain_admin_list/entry?{p}")

    def test_delete_pending_entry(self, auth_client):
        """Ghost entry (no matching credential) can be deleted."""
        ws_id = self._ws(auth_client, "r72-del-ghost")
        self._upload(auth_client, ws_id, "corp.local", "ghost_user")
        assert ("corp.local", "ghost_user") in self._pending(auth_client, ws_id)

        r = self._delete(auth_client, ws_id, "corp.local", "ghost_user")
        assert r.status_code == 200 and r.json()["ok"] is True

        assert ("corp.local", "ghost_user") not in self._pending(auth_client, ws_id)

    def test_delete_matched_entry(self, auth_client):
        """Matched entry (credential exists) can also be deleted from watchlist."""
        ws_id = self._ws(auth_client, "r72-del-matched")
        auth_client.post("/api/sync", json={
            "workspace": "r72-del-matched", "operator": "tester",
            "data": {
                "hosts": [], "auth_relations": [], "dpapi_secrets": [],
                "shares": [], "ssh_keys": [], "conf_checks_results": [],
                "directory_listings": [], "vuln_findings": [],
                "credentials": [{"proto": "SMB", "domain": "corp.local",
                                  "username": "da_one", "password": "Pass1",
                                  "credtype": "plaintext"}],
            },
        })
        self._upload(auth_client, ws_id, "corp.local", "da_one")

        # Verify entry is NOT in pending (it matched) — it's in DAL but not ghost
        pending = self._pending(auth_client, ws_id)
        assert ("corp.local", "da_one") not in pending

        r = self._delete(auth_client, ws_id, "corp.local", "da_one")
        assert r.status_code == 200

    def test_delete_404_for_missing_entry(self, auth_client):
        """Deleting a non-existent entry returns 404."""
        ws_id = self._ws(auth_client, "r72-404-test")
        r = self._delete(auth_client, ws_id, "corp.local", "nobody")
        assert r.status_code == 404

    def test_delete_is_case_insensitive(self, auth_client):
        """Entry uploaded as lowercase can be deleted with mixed case."""
        ws_id = self._ws(auth_client, "r72-case-test")
        self._upload(auth_client, ws_id, "corp.local", "CaseUser")
        # Upload lowercases; delete via original-case should still work
        r = self._delete(auth_client, ws_id, "CORP.LOCAL", "CaseUser")
        assert r.status_code == 200
        assert ("corp.local", "caseuser") not in self._pending(auth_client, ws_id)

    def test_delete_does_not_touch_admin_cred(self, auth_client):
        """Deleting a DAL entry does NOT automatically unmark admin_cred on matching credentials."""
        from collector.db import db_cursor
        ws_id = self._ws(auth_client, "r72-admcred-test")
        auth_client.post("/api/sync", json={
            "workspace": "r72-admcred-test", "operator": "tester",
            "data": {
                "hosts": [], "auth_relations": [], "dpapi_secrets": [],
                "shares": [], "ssh_keys": [], "conf_checks_results": [],
                "directory_listings": [], "vuln_findings": [],
                "credentials": [{"proto": "SMB", "domain": "corp.local",
                                  "username": "admuser", "password": "P@ss1",
                                  "credtype": "plaintext"}],
            },
        })
        self._upload(auth_client, ws_id, "corp.local", "admuser")

        with db_cursor() as cur:
            row = cur.execute(
                "SELECT admin_cred FROM credentials WHERE workspace_id=? AND username=?",
                (ws_id, "admuser")
            ).fetchone()
        assert row["admin_cred"] == 1, "watchlist must have marked admin_cred=1"

        self._delete(auth_client, ws_id, "corp.local", "admuser")

        with db_cursor() as cur:
            row = cur.execute(
                "SELECT admin_cred FROM credentials WHERE workspace_id=? AND username=?",
                (ws_id, "admuser")
            ).fetchone()
        assert row["admin_cred"] == 1, "admin_cred must remain unchanged after DAL entry deletion"


# ═══════════════════════════════════════════════════════════════════════════════
# R3.2 — Custom Import: matched/unrecognized header fields in response
# ═══════════════════════════════════════════════════════════════════════════════

class TestCustomImportHeaderFeedback:
    def _make_xlsx(self, headers, *data_rows):
        import openpyxl, io as _io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for r in data_rows:
            ws.append(list(r))
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _upload(self, auth_client, workspace_id, xlsx_bytes):
        return auth_client.post(
            "/api/toolbox/custom-import/upload",
            data={"workspace_id": str(workspace_id)},
            files={"file": ("test.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    def test_response_includes_header_fields(self, auth_client, workspace_id):
        """Both matched_headers and unrecognized_headers must always be present."""
        xlsx = self._make_xlsx(["Login", "Password"], ["alice", "P@ss1"])
        d = self._upload(auth_client, workspace_id, xlsx).json()
        assert "matched_headers" in d
        assert "unrecognized_headers" in d

    def test_known_columns_appear_in_matched(self, auth_client, workspace_id):
        xlsx = self._make_xlsx(
            ["Proto", "IP", "Login", "Password"],
            ["SMB", "10.33.0.1", "bob", "S3cr3t"],
        )
        d = self._upload(auth_client, workspace_id, xlsx).json()
        assert "Login" in d["matched_headers"]
        assert "Password" in d["matched_headers"]
        assert d["unrecognized_headers"] == []

    def test_wrong_columns_all_unrecognized(self, auth_client, workspace_id):
        xlsx = self._make_xlsx(
            ["Name", "Value", "Hash", "Tag"],
            ["DC01", "secret", "abc123", "foo"],
        )
        d = self._upload(auth_client, workspace_id, xlsx).json()
        assert d["matched_headers"] == []
        assert set(d["unrecognized_headers"]) == {"Name", "Value", "Hash", "Tag"}
        assert d["added"] == 0

    def test_mixed_columns_correct_split(self, auth_client, workspace_id):
        xlsx = self._make_xlsx(
            ["Login", "Password", "Unknown1", "Unknown2"],
            ["charlie", "pass", "x", "y"],
        )
        d = self._upload(auth_client, workspace_id, xlsx).json()
        assert "Login" in d["matched_headers"]
        assert "Password" in d["matched_headers"]
        assert "Unknown1" in d["unrecognized_headers"]
        assert "Unknown2" in d["unrecognized_headers"]


# ═══════════════════════════════════════════════════════════════════════════════
# R2.4 — Case-insensitive workspace name on creation
# ═══════════════════════════════════════════════════════════════════════════════

class TestWorkspaceNameCaseInsensitive:
    def _cleanup(self, ws_id):
        from collector.db import db_cursor
        with db_cursor() as cur:
            cur.execute("DELETE FROM workspaces WHERE id=?", (ws_id,))

    def test_uppercase_variant_rejected(self, auth_client):
        import uuid
        name = f"__ci_{uuid.uuid4().hex[:6]}__"
        ws_id = auth_client.post("/api/workspaces", json={"name": name}).json()["id"]
        try:
            r = auth_client.post("/api/workspaces", json={"name": name.upper()})
            assert r.status_code == 409
            assert "already exists" in r.json().get("detail", "").lower()
        finally:
            self._cleanup(ws_id)

    def test_lowercase_variant_rejected(self, auth_client):
        import uuid
        name = f"__CI_{uuid.uuid4().hex[:6]}__"
        ws_id = auth_client.post("/api/workspaces", json={"name": name}).json()["id"]
        try:
            r = auth_client.post("/api/workspaces", json={"name": name.lower()})
            assert r.status_code == 409
        finally:
            self._cleanup(ws_id)

    def test_mixed_case_variant_rejected(self, auth_client):
        import uuid
        name = f"CorpProject_{uuid.uuid4().hex[:6]}"
        ws_id = auth_client.post("/api/workspaces", json={"name": name}).json()["id"]
        try:
            r = auth_client.post("/api/workspaces", json={"name": name.swapcase()})
            assert r.status_code == 409
        finally:
            self._cleanup(ws_id)

    def test_different_name_still_succeeds(self, auth_client):
        import uuid
        a = f"__ci_a_{uuid.uuid4().hex[:6]}__"
        b = f"__ci_b_{uuid.uuid4().hex[:6]}__"
        ra = auth_client.post("/api/workspaces", json={"name": a})
        rb = auth_client.post("/api/workspaces", json={"name": b})
        assert ra.status_code == 200
        assert rb.status_code == 200
        from collector.db import db_cursor
        with db_cursor() as cur:
            cur.execute("DELETE FROM workspaces WHERE name IN (?,?)", (a, b))


# ═══════════════════════════════════════════════════════════════════════════════
# R2.2 — Duplicate hostname detection in /api/hosts
# ═══════════════════════════════════════════════════════════════════════════════

class TestDuplicateHostDetection:
    _HOST_EXTRA = {
        "signing": 1, "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0,
        "dc": 0, "nla": None, "signing_required": None, "channel_binding": None,
        "port": 445, "banner": None, "instances": None,
    }

    def _ws(self, auth_client, name):
        return auth_client.post("/api/workspaces", json={"name": name}).json()["id"]

    def _sync(self, auth_client, ws_name, hosts):
        return auth_client.post("/api/sync", json={
            "workspace": ws_name, "operator": "tester",
            "data": {
                "hosts": hosts,
                "credentials": [], "auth_relations": [], "dpapi_secrets": [],
                "shares": [], "ssh_keys": [], "conf_checks_results": [],
                "directory_listings": [], "vuln_findings": [],
            },
        })

    def _host(self, ip, hostname="", domain="corp.local"):
        return {"ip": ip, "hostname": hostname, "domain": domain, **self._HOST_EXTRA}

    def test_single_host_no_dup_flag(self, auth_client):
        ws_id = self._ws(auth_client, "r22-single")
        self._sync(auth_client, "r22-single", [self._host("10.22.0.1", "DC01")])
        rows = auth_client.get(f"/api/hosts?workspace_id={ws_id}").json()["rows"]
        assert len(rows) == 1
        assert rows[0]["_dup_hostname"] is False

    def test_two_hosts_same_hostname_both_flagged(self, auth_client):
        ws_id = self._ws(auth_client, "r22-same-hn")
        self._sync(auth_client, "r22-same-hn", [
            self._host("10.22.1.1", "DC01"),
            self._host("10.22.1.2", "DC01"),
        ])
        rows = sorted(
            auth_client.get(f"/api/hosts?workspace_id={ws_id}").json()["rows"],
            key=lambda r: r["ip"],
        )
        assert len(rows) == 2
        assert rows[0]["_dup_hostname"] is True
        assert rows[1]["_dup_hostname"] is True

    def test_dup_detection_case_insensitive(self, auth_client):
        ws_id = self._ws(auth_client, "r22-case-hn")
        self._sync(auth_client, "r22-case-hn", [
            self._host("10.22.2.1", "DC01"),
            self._host("10.22.2.2", "dc01"),
        ])
        rows = auth_client.get(f"/api/hosts?workspace_id={ws_id}").json()["rows"]
        assert all(r["_dup_hostname"] is True for r in rows)

    def test_unique_hostname_not_flagged(self, auth_client):
        ws_id = self._ws(auth_client, "r22-mixed")
        self._sync(auth_client, "r22-mixed", [
            self._host("10.22.3.1", "DC01"),
            self._host("10.22.3.2", "DC01"),
            self._host("10.22.3.3", "WEB01"),
        ])
        rows = {r["ip"]: r for r in
                auth_client.get(f"/api/hosts?workspace_id={ws_id}").json()["rows"]}
        assert rows["10.22.3.1"]["_dup_hostname"] is True
        assert rows["10.22.3.2"]["_dup_hostname"] is True
        assert rows["10.22.3.3"]["_dup_hostname"] is False

    def test_empty_hostname_never_flagged(self, auth_client):
        ws_id = self._ws(auth_client, "r22-empty-hn")
        self._sync(auth_client, "r22-empty-hn", [
            self._host("10.22.4.1", ""),
            self._host("10.22.4.2", ""),
        ])
        rows = auth_client.get(f"/api/hosts?workspace_id={ws_id}").json()["rows"]
        assert all(r["_dup_hostname"] is False for r in rows)

    def test_dup_flag_workspace_wide_despite_search_filter(self, auth_client):
        """Search filter must not affect dup detection — detection is always workspace-wide."""
        ws_id = self._ws(auth_client, "r22-search-scope")
        self._sync(auth_client, "r22-search-scope", [
            self._host("10.22.5.1", "DC01"),
            self._host("10.22.5.2", "DC01"),
        ])
        rows = auth_client.get(
            f"/api/hosts?workspace_id={ws_id}&search=10.22.5.1"
        ).json()["rows"]
        assert len(rows) == 1
        assert rows[0]["_dup_hostname"] is True
