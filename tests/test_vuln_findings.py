"""
Phase 1 — vuln_findings table + sync UPSERT (tri-state).

Tri-state is_vulnerable: 1=vulnerable, 0=checked-clean, NULL=could-not-check.
UPSERT priority on (workspace_id, ip, vuln_name): 1 > 0 > NULL; ties → latest timestamp.
"""

from io import BytesIO

import openpyxl
import pytest

from collector.db import db_cursor


def _sync_findings(auth_client, ws_name, findings):
    """Push a sync payload carrying only vuln_findings."""
    r = auth_client.post("/api/sync", json={
        "workspace": ws_name, "operator": "tester",
        "data": {"vuln_findings": findings},
    })
    assert r.status_code == 200
    return r


def _finding(ip, vuln_name, is_vulnerable, details="", hostname="H", domain="corp.local"):
    return {
        "ip": ip, "hostname": hostname, "domain": domain,
        "protocol": "smb", "port": 445,
        "vuln_name": vuln_name, "is_vulnerable": is_vulnerable, "details": details,
    }


def _read(ws_id, ip, vuln_name):
    with db_cursor() as cur:
        return cur.execute(
            "SELECT is_vulnerable, details FROM vuln_findings"
            " WHERE workspace_id=? AND ip=? AND vuln_name=?",
            (ws_id, ip, vuln_name),
        ).fetchone()


@pytest.fixture
def ws(auth_client):
    """Fresh workspace per test."""
    import uuid
    name = f"vf-{uuid.uuid4().hex[:8]}"
    wid = auth_client.post("/api/workspaces", json={"name": name}).json()["id"]
    return name, wid


def test_sync_inserts_vuln_finding(auth_client, ws):
    name, wid = ws
    _sync_findings(auth_client, name, [_finding("10.0.0.1", "zerologon", 1, "vuln!")])
    row = _read(wid, "10.0.0.1", "zerologon")
    assert row is not None, "vuln_finding was not stored"
    assert row["is_vulnerable"] == 1
    assert row["details"] == "vuln!"


def test_dedup_single_row_per_ip_vuln(auth_client, ws):
    name, wid = ws
    _sync_findings(auth_client, name, [_finding("10.0.0.2", "webdav", 0)])
    _sync_findings(auth_client, name, [_finding("10.0.0.2", "webdav", 0)])
    with db_cursor() as cur:
        n = cur.execute(
            "SELECT COUNT(*) FROM vuln_findings WHERE workspace_id=? AND ip=? AND vuln_name=?",
            (wid, "10.0.0.2", "webdav"),
        ).fetchone()[0]
    assert n == 1, f"expected one row per (ip, vuln_name), got {n}"


def test_vulnerable_wins_over_clean(auth_client, ws):
    name, wid = ws
    _sync_findings(auth_client, name, [_finding("10.0.0.3", "ms17_010", 0, "clean")])
    _sync_findings(auth_client, name, [_finding("10.0.0.3", "ms17_010", 1, "vuln")])
    assert _read(wid, "10.0.0.3", "ms17_010")["is_vulnerable"] == 1


def test_vulnerable_not_overwritten_by_clean(auth_client, ws):
    name, wid = ws
    _sync_findings(auth_client, name, [_finding("10.0.0.4", "smbghost", 1, "vuln")])
    _sync_findings(auth_client, name, [_finding("10.0.0.4", "smbghost", 0, "clean")])
    assert _read(wid, "10.0.0.4", "smbghost")["is_vulnerable"] == 1, \
        "vulnerable-wins: a later clean result must not downgrade a vulnerable finding"


def test_clean_beats_null(auth_client, ws):
    name, wid = ws
    _sync_findings(auth_client, name, [_finding("10.0.0.5", "printnightmare", None, "error")])
    assert _read(wid, "10.0.0.5", "printnightmare")["is_vulnerable"] is None
    _sync_findings(auth_client, name, [_finding("10.0.0.5", "printnightmare", 0, "clean")])
    assert _read(wid, "10.0.0.5", "printnightmare")["is_vulnerable"] == 0, \
        "a real checked-clean (0) must replace a could-not-check (NULL)"


def test_null_does_not_overwrite_clean_or_vuln(auth_client, ws):
    name, wid = ws
    # clean then null → stays clean
    _sync_findings(auth_client, name, [_finding("10.0.0.6", "uac", 0, "clean")])
    _sync_findings(auth_client, name, [_finding("10.0.0.6", "uac", None, "error")])
    assert _read(wid, "10.0.0.6", "uac")["is_vulnerable"] == 0, \
        "could-not-check (NULL) must not overwrite a checked-clean (0)"
    # vuln then null → stays vuln
    _sync_findings(auth_client, name, [_finding("10.0.0.6", "wdigest", 1, "vuln")])
    _sync_findings(auth_client, name, [_finding("10.0.0.6", "wdigest", None, "error")])
    assert _read(wid, "10.0.0.6", "wdigest")["is_vulnerable"] == 1, \
        "could-not-check (NULL) must not overwrite a vulnerable (1)"


def test_null_finding_stored_as_null(auth_client, ws):
    name, wid = ws
    _sync_findings(auth_client, name, [_finding("10.0.0.7", "ntlmv1", None, "registry error")])
    row = _read(wid, "10.0.0.7", "ntlmv1")
    assert row is not None
    assert row["is_vulnerable"] is None, "could-not-check must persist as NULL, not 0"


# ───────────────────────────────────────────────────────────────────────────
# Phase 3 — get_vulns pivot (aggregate-per-host, merge, tri-state)
# ───────────────────────────────────────────────────────────────────────────

def _host(ip, **vulns):
    h = {"ip": ip, "hostname": "H", "domain": "corp.local", "os": "Win",
         "dc": 0, "smbv1": None, "signing": None, "spooler": None,
         "zerologon": None, "petitpotam": None, "nla": None,
         "signing_required": None, "channel_binding": None,
         "port": 445, "banner": None, "instances": None}
    h.update(vulns)
    return h


def _sync(auth_client, ws_name, hosts=None, findings=None):
    r = auth_client.post("/api/sync", json={
        "workspace": ws_name, "operator": "tester",
        "data": {"hosts": hosts or [], "vuln_findings": findings or []},
    })
    assert r.status_code == 200


def _vulns(auth_client, wid, **params):
    q = "&".join(f"{k}={v}" for k, v in params.items())
    r = auth_client.get(f"/api/vulns?workspace_id={wid}&{q}")
    assert r.status_code == 200
    return r.json()["rows"]


def _row(rows, ip):
    return next((r for r in rows if r["ip"] == ip), None)


def test_pivot_host_column_smbv1(auth_client, ws):
    name, wid = ws
    _sync(auth_client, name, hosts=[_host("10.1.0.1", smbv1=1)])
    rows = _vulns(auth_client, wid)
    r = _row(rows, "10.1.0.1")
    assert r is not None, "vulnerable host must appear in ALL"
    assert r["smbv1"] == 1


def test_pivot_signing_inverted(auth_client, ws):
    name, wid = ws
    # signing=0 means SMB signing OFF → vulnerable; signing=1 → clean
    _sync(auth_client, name, hosts=[_host("10.1.0.2", signing=0), _host("10.1.0.3", signing=1)])
    rows = _vulns(auth_client, wid)
    assert _row(rows, "10.1.0.2")["signing"] == 1, "signing=0 is the vulnerable state"
    # 10.1.0.3 is fully clean → must NOT appear in ALL
    assert _row(rows, "10.1.0.3") is None


def test_pivot_clean_host_excluded_from_all(auth_client, ws):
    name, wid = ws
    _sync(auth_client, name, hosts=[_host("10.1.0.4", smbv1=0, signing=1, spooler=0)])
    assert _row(_vulns(auth_client, wid), "10.1.0.4") is None


def test_pivot_finding_based(auth_client, ws):
    name, wid = ws
    _sync(auth_client, name, hosts=[_host("10.1.0.5")],
          findings=[_finding("10.1.0.5", "ms17_010", 1, "vuln")])
    r = _row(_vulns(auth_client, wid), "10.1.0.5")
    assert r is not None and r["ms17_010"] == 1


def test_pivot_finding_tri_state_null(auth_client, ws):
    name, wid = ws
    _sync(auth_client, name, hosts=[_host("10.1.0.6", smbv1=1)],
          findings=[_finding("10.1.0.6", "ms17_010", None, "error")])
    r = _row(_vulns(auth_client, wid), "10.1.0.6")
    # smbv1=1 keeps the host in ALL; ms17_010 could-not-check must be null, NOT 0
    assert r["ms17_010"] is None, "could-not-check must surface as null, never as clean 0"


def test_pivot_merge_petitpotam_from_finding(auth_client, ws):
    name, wid = ws
    # host petitpotam column clean (0), but a finding says vulnerable → merged = vulnerable
    _sync(auth_client, name, hosts=[_host("10.1.0.7", petitpotam=0)],
          findings=[_finding("10.1.0.7", "petitpotam", 1, "pipe")])
    r = _row(_vulns(auth_client, wid), "10.1.0.7")
    assert r is not None and r["petitpotam"] == 1, "merge: finding vulnerable wins"


def test_pivot_merge_zerologon_from_hostcol(auth_client, ws):
    name, wid = ws
    # legacy host-column zerologon=1, no finding → merged = vulnerable
    _sync(auth_client, name, hosts=[_host("10.1.0.8", zerologon=1)])
    r = _row(_vulns(auth_client, wid), "10.1.0.8")
    assert r is not None and r["zerologon"] == 1


def test_pivot_filter_by_vuln_slug(auth_client, ws):
    name, wid = ws
    _sync(auth_client, name,
          hosts=[_host("10.1.0.9", smbv1=1), _host("10.1.0.10")],
          findings=[_finding("10.1.0.10", "webdav", 1, "v")])
    rows = _vulns(auth_client, wid, vuln="webdav")
    ips = {r["ip"] for r in rows}
    assert "10.1.0.10" in ips and "10.1.0.9" not in ips, "vuln filter keeps only matching hosts"


# ───────────────────────────────────────────────────────────────────────────
# Phase 4 — XLSX export mirrors the VULNS view (VULN_COLUMNS)
# ───────────────────────────────────────────────────────────────────────────

def test_export_vulns_columns_and_cells(auth_client, ws):
    from collector.core.constants import VULN_COLUMNS
    name, wid = ws
    _sync(auth_client, name, hosts=[_host("10.2.0.1", smbv1=1, signing=1)],
          findings=[_finding("10.2.0.1", "ms17_010", 1, "v"),
                    _finding("10.2.0.1", "uac", 0, "clean")])
    r = auth_client.get(f"/api/export/xlsx?workspace_id={wid}&view=vulns")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    ws_xl = wb.active
    headers = [ws_xl.cell(1, c).value for c in range(1, 5 + len(VULN_COLUMNS))]
    # IP..OS then every VULN_COLUMNS label, in order
    assert headers[:4] == ["IP", "Hostname", "Domain", "OS"]
    assert headers[4:4 + len(VULN_COLUMNS)] == [c["label"] for c in VULN_COLUMNS]
    # Row values: find the data row
    data_row = [ws_xl.cell(2, c).value for c in range(1, 5 + len(VULN_COLUMNS))]
    by_label = dict(zip(headers, data_row))
    assert by_label["SMBv1"] == "YES"
    assert by_label["MS17-010"] == "YES"
    assert by_label["UAC"] == "no"            # checked-clean → "no"
    assert by_label["Signing OFF"] == "no"    # signing=1 → clean (inverted) → "no"
    assert by_label["WebDAV"] in ("", None)   # no finding → blank cell


# ═══════════════════════════════════════════════════════════════════════════════
# Phase 5 — R6.1: manual vuln overrides (user value > sync value)
# ═══════════════════════════════════════════════════════════════════════════════

def _set_override(auth_client, wid, ip, vuln_name, is_vulnerable):
    return auth_client.post("/api/vulns/set_override", json={
        "workspace_id": wid, "ip": ip,
        "vuln_name": vuln_name, "is_vulnerable": is_vulnerable,
    })


def test_override_beats_sync_yes(auth_client, ws):
    """Sync marks host as YES (1); user sets NO (0) → pivot returns 0."""
    name, wid = ws
    _sync(auth_client, name, findings=[_finding("10.5.0.1", "ms17_010", 1)])
    assert _row(_vulns(auth_client, wid, vuln="ms17_010"), "10.5.0.1")["ms17_010"] == 1

    r = _set_override(auth_client, wid, "10.5.0.1", "ms17_010", 0)
    assert r.status_code == 200

    rows = auth_client.get(f"/api/vulns?workspace_id={wid}&vuln=all").json()["rows"]
    found = _row(rows, "10.5.0.1")
    # After override=0, host is no longer vulnerable → won't appear in ALL filter
    # but direct value check via GET /api/vulns with no filter (limit=0)
    all_rows = auth_client.get(f"/api/vulns?workspace_id={wid}&limit=0").json()["rows"]
    # Host is gone from ALL (no longer vulnerable after override); use unfiltered endpoint
    # Check via direct vuln pivot without filter
    r2 = auth_client.get(f"/api/vulns?workspace_id={wid}&vuln=ms17_010")
    assert r2.status_code == 200
    assert _row(r2.json()["rows"], "10.5.0.1") is None, "override=0 must remove host from YES filter"


def test_override_resync_does_not_overwrite(auth_client, ws):
    """After user sets NO (0), re-sync with YES (1) must NOT restore YES."""
    name, wid = ws
    _sync(auth_client, name, findings=[_finding("10.5.0.2", "ms17_010", 1)])
    _set_override(auth_client, wid, "10.5.0.2", "ms17_010", 0)

    # Re-sync with YES again
    _sync(auth_client, name, findings=[_finding("10.5.0.2", "ms17_010", 1)])

    r = auth_client.get(f"/api/vulns?workspace_id={wid}&vuln=ms17_010")
    assert _row(r.json()["rows"], "10.5.0.2") is None, (
        "re-sync must not overwrite user override"
    )


def test_override_none_explicit(auth_client, ws):
    """User sets NONE (null) explicitly → pivot returns null (not sync YES)."""
    name, wid = ws
    _sync(auth_client, name, findings=[_finding("10.5.0.3", "ms17_010", 1)])
    _set_override(auth_client, wid, "10.5.0.3", "ms17_010", None)

    r = auth_client.get(f"/api/vulns?workspace_id={wid}&vuln=ms17_010")
    assert _row(r.json()["rows"], "10.5.0.3") is None, "override=null must suppress YES"

    # Check unfiltered to confirm value is null (not 1)
    all_r = auth_client.get(f"/api/vulns?workspace_id={wid}&limit=0")
    assert all_r.status_code == 200
    # Host should NOT appear in ALL (no vulns after override=null)
    assert _row(all_r.json()["rows"], "10.5.0.3") is None


def test_override_host_source_vuln(auth_client, ws):
    """Override works on host-source vulns (smbv1) — not just finding-source."""
    name, wid = ws
    _sync(auth_client, name, hosts=[_host("10.5.0.4", smbv1=1)])
    rows = _vulns(auth_client, wid)
    assert _row(rows, "10.5.0.4")["smbv1"] == 1

    _set_override(auth_client, wid, "10.5.0.4", "smbv1", 0)

    rows2 = auth_client.get(f"/api/vulns?workspace_id={wid}&limit=0").json()["rows"]
    r = _row(rows2, "10.5.0.4")
    assert r is None or r.get("smbv1") == 0, "user override must beat host-column source"


def test_override_can_upgrade_none_to_yes(auth_client, ws):
    """User can set YES (1) even if sync has no finding (null → 1)."""
    name, wid = ws
    _sync(auth_client, name, hosts=[_host("10.5.0.5")])
    _set_override(auth_client, wid, "10.5.0.5", "ms17_010", 1)

    r = auth_client.get(f"/api/vulns?workspace_id={wid}&vuln=ms17_010")
    assert _row(r.json()["rows"], "10.5.0.5") is not None, "override=1 must make host appear"
    assert _row(r.json()["rows"], "10.5.0.5")["ms17_010"] == 1


def test_override_cycle(auth_client, ws):
    """Full 3-step cycle: 1 → 0 → null → 1 (all three stored and returned correctly).
    Host also has smbv1=1 (host-column) so it stays visible in ALL filter at every step."""
    name, wid = ws
    # smbv1=1 keeps the host in ALL throughout; ms17_010 is the cycled slug
    _sync(auth_client, name,
          hosts=[_host("10.5.0.6", smbv1=1)],
          findings=[_finding("10.5.0.6", "ms17_010", 1)])

    def get_val():
        all_r = auth_client.get(f"/api/vulns?workspace_id={wid}&limit=0")
        row = _row(all_r.json()["rows"], "10.5.0.6")
        return row["ms17_010"] if row else "HOST_MISSING"

    # Sync: YES
    assert get_val() == 1

    # Click 1: YES → NO
    _set_override(auth_client, wid, "10.5.0.6", "ms17_010", 0)
    assert get_val() == 0

    # Click 2: NO → NONE
    _set_override(auth_client, wid, "10.5.0.6", "ms17_010", None)
    assert get_val() is None

    # Click 3: NONE → YES
    _set_override(auth_client, wid, "10.5.0.6", "ms17_010", 1)
    assert get_val() == 1


# ═══════════════════════════════════════════════════════════════════════════════
# Phase 6 — show_all: manage mode sees all hosts even with no vulnerable slugs
# ═══════════════════════════════════════════════════════════════════════════════

def test_show_all_includes_non_vulnerable_host(auth_client, ws):
    """show_all=1 returns hosts even when all their vuln values are 0 or null."""
    name, wid = ws
    _sync(auth_client, name, hosts=[_host("10.6.0.1", smbv1=1)])

    # Without show_all: host appears (smbv1=1 is vulnerable)
    rows = auth_client.get(f"/api/vulns?workspace_id={wid}").json()["rows"]
    assert _row(rows, "10.6.0.1") is not None

    # Override smbv1 to 0 — now host has no vulnerable slugs
    _set_override(auth_client, wid, "10.6.0.1", "smbv1", 0)

    # Without show_all: host disappears (no vulnerable slugs)
    rows = auth_client.get(f"/api/vulns?workspace_id={wid}").json()["rows"]
    assert _row(rows, "10.6.0.1") is None

    # With show_all=1: host is visible despite having no vulnerable slugs
    rows = auth_client.get(f"/api/vulns?workspace_id={wid}&show_all=1").json()["rows"]
    r = _row(rows, "10.6.0.1")
    assert r is not None, "show_all must include hosts with no vulnerable slugs"
    assert r["smbv1"] == 0


def test_clear_all_overrides(auth_client, ws):
    """DELETE /api/vulns/overrides clears all user overrides for the workspace."""
    name, wid = ws
    _sync(auth_client, name, findings=[
        _finding("10.7.0.1", "ms17_010", 1),
        _finding("10.7.0.2", "nopac", 1),
    ])
    _set_override(auth_client, wid, "10.7.0.1", "ms17_010", 0)
    _set_override(auth_client, wid, "10.7.0.2", "nopac", 0)

    r = auth_client.delete(f"/api/vulns/overrides?workspace_id={wid}")
    assert r.status_code == 200
    assert r.json()["ok"] is True
    assert r.json()["deleted"] == 2

    # After clear, sync values are back
    rows = auth_client.get(f"/api/vulns?workspace_id={wid}").json()["rows"]
    assert _row(rows, "10.7.0.1")["ms17_010"] == 1
    assert _row(rows, "10.7.0.2")["nopac"] == 1


def test_show_all_does_not_break_vuln_filter(auth_client, ws):
    """show_all=1 combined with vuln= filter still applies the slug filter."""
    name, wid = ws
    _sync(auth_client, name,
          hosts=[_host("10.6.0.2", smbv1=1)],
          findings=[_finding("10.6.0.2", "ms17_010", 1)])

    # vuln=ms17_010 + show_all: only hosts with ms17_010=1 appear
    rows = auth_client.get(f"/api/vulns?workspace_id={wid}&vuln=ms17_010&show_all=1").json()["rows"]
    assert _row(rows, "10.6.0.2") is not None

    # Override ms17_010 to 0 — now ms17_010 filter should exclude the host
    _set_override(auth_client, wid, "10.6.0.2", "ms17_010", 0)
    rows = auth_client.get(f"/api/vulns?workspace_id={wid}&vuln=ms17_010&show_all=1").json()["rows"]
    assert _row(rows, "10.6.0.2") is None, "vuln= filter must still apply even with show_all=1"
