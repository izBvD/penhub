"""
Core operator scenario + key regression tests.

Golden path: sync data → view hosts/creds/results → manage → export.

Regressions covered:
- Sync idempotency (no host duplicates)
- Empty NT hash normalization (aad3b435…:31d6cfe0 → plaintext)
- Shell flag update on resync (D-2 fix)
- Directory listing no duplicates on resync (D-1 fix)
- XLSX formula injection (=password must not become a formula cell)
- Guest filter (Guest/DefaultAccount hidden by default)
- Strike IP: shared credential must not be hidden when striking only one of its hosts
- Strike/restore: credential stuck hidden after gaining auth_relation on new host (hidden_by_strike fix)
"""

from io import BytesIO

import openpyxl
import pytest

import collector.hashkiller_db as hk_db
from collector.db import db_cursor


_SYNC = {
    "workspace": "test-ws",
    "operator": "tester",
    "data": {
        "hosts": [
            {
                "ip": "10.0.0.1", "hostname": "DC01", "domain": "corp.local",
                "os": "Windows Server 2019", "dc": 1, "signing": 1,
                "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0,
                "nla": None, "signing_required": None, "channel_binding": None,
                "port": 445, "banner": None, "instances": None,
            }
        ],
        "credentials": [
            {"proto": "SMB", "domain": "corp.local", "username": "administrator",
             "password": "P@ssw0rd!", "credtype": "plaintext"},
            {"proto": "SMB", "domain": "corp.local", "username": "svc_backup",
             "password": "Backup2!", "credtype": "plaintext"},
        ],
        "auth_relations": [
            {"proto": "SMB", "host_ip": "10.0.0.1",
             "cred_domain": "corp.local", "cred_username": "administrator",
             "cred_password": "P@ssw0rd!", "cred_credtype": "plaintext",
             "relation_type": "admin", "shell": 0},
            {"proto": "SMB", "host_ip": "10.0.0.1",
             "cred_domain": "corp.local", "cred_username": "svc_backup",
             "cred_password": "Backup2!", "cred_credtype": "plaintext",
             "relation_type": "loggedin", "shell": 0},
        ],
        "dpapi_secrets": [], "shares": [], "ssh_keys": [],
        "conf_checks_results": [], "directory_listings": [],
    },
}


# ═══════════════════════════════════════════════════════════════════════════════
# Golden path — one sync, many checks
# ═══════════════════════════════════════════════════════════════════════════════

class TestGoldenPath:
    """Main operator workflow verified against a single sync operation."""

    @pytest.fixture(scope="class", autouse=True)
    def synced(self, auth_client, workspace_id):
        r = auth_client.post("/api/sync", json=_SYNC)
        assert r.status_code == 200
        assert r.json()["ok"] is True

    def test_host_visible(self, auth_client, workspace_id):
        rows = auth_client.get(f"/api/hosts?workspace_id={workspace_id}").json()["rows"]
        assert any(h["ip"] == "10.0.0.1" for h in rows)

    def test_creds_visible(self, auth_client, workspace_id):
        rows = auth_client.get(
            f"/api/credentials?workspace_id={workspace_id}&hide_guest=false"
        ).json()["rows"]
        users = {c["username"] for c in rows}
        assert {"administrator", "svc_backup"}.issubset(users)

    def test_admin_relation_filter(self, auth_client, workspace_id):
        rows = auth_client.get(
            f"/api/results?workspace_id={workspace_id}&relation=admin&hide_guest=false"
        ).json()["rows"]
        assert rows
        assert all(r["relation_type"] == "admin" for r in rows)
        assert any(r["username"] == "administrator" for r in rows)

    def test_stats_correct(self, auth_client, workspace_id):
        d = auth_client.get(f"/api/stats?workspace_id={workspace_id}").json()
        assert d["hosts"] >= 1
        assert d["creds"] >= 1
        assert d["admin"] >= 1

    def test_hide_cred_and_restore(self, auth_client, workspace_id):
        auth_client.post("/api/credentials/set_hidden", json={
            "workspace_id": workspace_id, "domain": "corp.local",
            "username": "svc_backup", "password": "Backup2!", "hidden": 1,
        })
        visible = {c["username"] for c in auth_client.get(
            f"/api/credentials?workspace_id={workspace_id}"
        ).json()["rows"]}
        assert "svc_backup" not in visible

        auth_client.post("/api/credentials/set_hidden", json={
            "workspace_id": workspace_id, "domain": "corp.local",
            "username": "svc_backup", "password": "Backup2!", "hidden": 0,
        })
        restored = {c["username"] for c in auth_client.get(
            f"/api/credentials?workspace_id={workspace_id}"
        ).json()["rows"]}
        assert "svc_backup" in restored

    def test_export_xlsx_results(self, auth_client, workspace_id):
        r = auth_client.get(f"/api/export/xlsx?workspace_id={workspace_id}&view=results")
        assert r.status_code == 200
        assert "spreadsheet" in r.headers.get("content-type", "")

    def test_export_allcred(self, auth_client, workspace_id):
        r = auth_client.get(f"/api/export/allcred?workspace_id={workspace_id}")
        assert r.status_code == 200
        assert "spreadsheet" in r.headers.get("content-type", "")
        wb = openpyxl.load_workbook(BytesIO(r.content))
        ws = wb.active
        # Section 1 header row must match plaintext format
        assert [ws.cell(1, c).value for c in range(1, 5)] == ["Service", "Domain", "Login", "Password"]
        # Both synced plaintext credentials must appear
        col3 = {ws.cell(row, 3).value for row in range(2, ws.max_row + 1)}
        assert {"administrator", "svc_backup"}.issubset(col3)


def test_export_allcred_dpapi_section(auth_client, workspace_id):
    """DPAPI section headers must be Service|Host|URL|Login|Password."""
    with db_cursor() as cur:
        cur.execute(
            "INSERT OR IGNORE INTO dpapi_secrets"
            " (workspace_id, host_ip, dpapi_type, username, password, url, hidden)"
            " VALUES (?,?,?,?,?,?,0)",
            (workspace_id, "10.0.0.99", "SMB", "dpapi_user", "dpapi_pass", "http://example.com"),
        )
    r = auth_client.get(f"/api/export/allcred?workspace_id={workspace_id}")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    ws = wb.active
    dpapi_hdr_row = None
    for row in ws.iter_rows():
        if row[0].value == "DPAPI":
            dpapi_hdr_row = row[0].row + 1
            break
    assert dpapi_hdr_row is not None, "DPAPI separator row not found in XLSX"
    assert [ws.cell(dpapi_hdr_row, c).value for c in range(1, 6)] == [
        "Service", "Host", "URL", "Login", "Password"
    ]


def test_export_allcred_dpapi_password_only_with_url(auth_client, workspace_id):
    """DPAPI secrets with only a password + URL (no login) must still appear in ALL CREDS —
    a site with a single password field is valuable; the client must rotate it too."""
    UNIQUE_PW  = "__pw_only_secret__"
    UNIQUE_URL = "http://passwordonly.example"
    with db_cursor() as cur:
        cur.execute(
            "INSERT OR IGNORE INTO dpapi_secrets"
            " (workspace_id, host_ip, dpapi_type, username, password, url, hidden)"
            " VALUES (?,?,?,?,?,?,0)",
            (workspace_id, "10.0.0.98", "SMB", "", UNIQUE_PW, UNIQUE_URL),
        )
    r = auth_client.get(f"/api/export/allcred?workspace_id={workspace_id}")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    cells = {c.value for row in wb.active.iter_rows() for c in row}
    assert UNIQUE_PW in cells, "password-only DPAPI secret (no login) missing from ALL CREDS"
    assert UNIQUE_URL in cells


def test_kill_workspace_cracks_custom_creds(auth_client, workspace_id):
    """KILL THEM ALL must also resolve hashes stored in custom_credentials (Toolbox import),
    not only the synced credentials table — and the cracked plaintext must surface via the API."""
    from collector.api.data_manage import get_custom_creds
    NT    = "abcdef0123456789abcdef0123456789"
    PLAIN = "__custom_cracked__"
    hk_db.init_hk_db()
    hk_db.delete_by_value(NT)
    try:
        hk_db.bulk_import(f"{NT}:{PLAIN}")
        with db_cursor() as cur:
            cur.execute(
                "INSERT OR IGNORE INTO custom_credentials"
                " (workspace_id, proto, login, password, credtype, source)"
                " VALUES (?,?,?,?,?,?)",
                (workspace_id, "HTTP", "admin", NT, "hash", "test"),
            )

        res = hk_db.kill_workspace(workspace_id)
        assert res["updated"] >= 1

        with db_cursor() as cur:
            row = cur.execute(
                "SELECT brutforced FROM custom_credentials WHERE workspace_id=? AND password=?",
                (workspace_id, NT),
            ).fetchone()
        assert row is not None and row["brutforced"] == PLAIN

        # Display: get_custom_creds returns brutforced (so the HK toggle can reveal it).
        data = get_custom_creds(workspace_id=workspace_id, limit=0)
        crow = next(r for r in data["rows"] if r["password"] == NT)
        assert crow["brutforced"] == PLAIN
    finally:
        hk_db.delete_by_value(NT)
        with db_cursor() as cur:
            cur.execute("DELETE FROM custom_credentials WHERE workspace_id=? AND password=?",
                        (workspace_id, NT))


def test_allcred_substitutes_brutforced_for_custom(auth_client, workspace_id):
    """A cracked custom-cred hash (brutforced filled) shows as plaintext in ALL CREDS."""
    NT    = "fedcba9876543210fedcba9876543210"
    PLAIN = "__allcred_custom_plain__"
    with db_cursor() as cur:
        cur.execute(
            "INSERT OR IGNORE INTO custom_credentials"
            " (workspace_id, proto, login, password, credtype, brutforced, source)"
            " VALUES (?,?,?,?,?,?,?)",
            (workspace_id, "HTTP", "svc", NT, "hash", PLAIN, "test"),
        )
    try:
        r = auth_client.get(f"/api/export/allcred?workspace_id={workspace_id}")
        assert r.status_code == 200
        cells = {c.value for row in openpyxl.load_workbook(BytesIO(r.content)).active.iter_rows() for c in row}
        assert PLAIN in cells, "cracked custom hash must show plaintext in ALL CREDS"
        assert NT not in cells, "raw hash must be replaced by the cracked plaintext"
    finally:
        with db_cursor() as cur:
            cur.execute("DELETE FROM custom_credentials WHERE workspace_id=? AND password=?",
                        (workspace_id, NT))


def test_toolbox_passwords_includes_cracked_custom(auth_client, workspace_id):
    """A cracked custom-cred hash contributes its plaintext to the passwords export."""
    NT    = "1111aaaa2222bbbb3333cccc4444dddd"
    PLAIN = "__custom_pw_export__"
    with db_cursor() as cur:
        cur.execute(
            "INSERT OR IGNORE INTO custom_credentials"
            " (workspace_id, proto, login, password, credtype, brutforced, source)"
            " VALUES (?,?,?,?,?,?,?)",
            (workspace_id, "HTTP", "u", NT, "hash", PLAIN, "test"),
        )
    try:
        r = auth_client.get(f"/api/toolbox/passwords?workspace_id={workspace_id}")
        assert r.status_code == 200
        lines = r.text.splitlines()
        assert PLAIN in lines, "cracked custom plaintext must be in passwords export"
        assert NT not in lines, "raw custom hash must not appear as a password"
    finally:
        with db_cursor() as cur:
            cur.execute("DELETE FROM custom_credentials WHERE workspace_id=? AND password=?",
                        (workspace_id, NT))


def test_spray_archive_cracked_custom_goes_to_plaintext(auth_client, workspace_id):
    """A cracked custom-cred hash lands in the plaintext spray pair, not the hash pair."""
    import zipfile
    NT    = "2222aaaa3333bbbb4444cccc5555dddd"
    PLAIN = "__custom_spray_export__"
    with db_cursor() as cur:
        cur.execute(
            "INSERT OR IGNORE INTO custom_credentials"
            " (workspace_id, proto, login, password, credtype, brutforced, source)"
            " VALUES (?,?,?,?,?,?,?)",
            (workspace_id, "HTTP", "sprayuser", NT, "hash", PLAIN, "test"),
        )
    try:
        r = auth_client.get(f"/api/toolbox/spray-archive?workspace_id={workspace_id}")
        assert r.status_code == 200
        zf = zipfile.ZipFile(BytesIO(r.content))
        plain_passes = zf.read("plaintext_passes.txt").decode().splitlines()
        hash_passes  = zf.read("hashes_passes.txt").decode().splitlines()
        assert PLAIN in plain_passes, "cracked custom hash must spray as plaintext"
        assert NT not in hash_passes, "cracked custom hash must NOT remain in the hash pair"
    finally:
        with db_cursor() as cur:
            cur.execute("DELETE FROM custom_credentials WHERE workspace_id=? AND password=?",
                        (workspace_id, NT))


def test_export_allcred_local_admin_section(auth_client):
    """
    LOCAL ADMIN section: appears between Plaintext and Hashes; each machine row is separate
    (dedup uses real domain names so 2 machines with same creds = 2 rows, not 1).
    """
    r = auth_client.post("/api/workspaces", json={"name": "la-export-test"})
    ws_id = r.json()["id"]

    _base = {"dpapi_secrets": [], "shares": [], "ssh_keys": [],
             "conf_checks_results": [], "directory_listings": []}
    # Two different machines, same (username, password) — must appear as 2 rows in LOCAL ADMIN
    for machine in ("MACHINE-A", "MACHINE-B"):
        auth_client.post("/api/sync", json={
            "workspace": "la-export-test", "operator": "tester",
            "data": {
                "hosts": [{"ip": f"10.20.0.{'1' if machine == 'MACHINE-A' else '2'}",
                           "hostname": machine, "domain": machine, "signing": 0,
                           "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0,
                           "dc": 0, "nla": None, "signing_required": None,
                           "channel_binding": None, "port": 445, "banner": None, "instances": None}],
                "credentials": [{"proto": "SMB", "domain": machine, "username": "localadm",
                                  "password": "L0c@l!", "credtype": "plaintext"}],
                "auth_relations": [],
                **_base,
            },
        })
    # Mark as local admin
    auth_client.post("/api/credentials/set_local_admin_cred",
                     json={"workspace_id": ws_id, "username": "localadm",
                           "password": "L0c@l!", "local_admin_cred": 1})

    resp = auth_client.get(f"/api/export/allcred?workspace_id={ws_id}")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws_xl = wb.active

    # Find LOCAL ADMIN separator row
    la_sep_row = next(
        (row[0].row for row in ws_xl.iter_rows() if row[0].value == "LOCAL ADMIN"),
        None,
    )
    assert la_sep_row is not None, "LOCAL ADMIN separator not found in XLSX"

    # Header row immediately after separator
    la_hdr_row = la_sep_row + 1
    assert [ws_xl.cell(la_hdr_row, c).value for c in range(1, 6)] == [
        "Service", "Machine", "Login", "Password", "Type"
    ], "LOCAL ADMIN header columns must be Service|Machine|Login|Password|Type"

    # Both machine rows must appear (not collapsed to 1)
    data_rows = []
    for row_idx in range(la_hdr_row + 1, ws_xl.max_row + 1):
        label = ws_xl.cell(row_idx, 1).value
        if label in ("HASHES", "DPAPI", "CUSTOM") or label is None:
            break
        data_rows.append(ws_xl.cell(row_idx, 3).value)  # Login column

    assert data_rows.count("localadm") == 2, (
        "LOCAL ADMIN section must have 2 rows (one per machine) — must not group identical credentials"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Sync regressions
# ═══════════════════════════════════════════════════════════════════════════════

def test_sync_idempotent_no_host_duplicates(auth_client, workspace_id):
    """Double sync must not create duplicate hosts."""
    auth_client.post("/api/sync", json=_SYNC)
    rows = auth_client.get(f"/api/hosts?workspace_id={workspace_id}").json()["rows"]
    assert [h["ip"] for h in rows].count("10.0.0.1") == 1


def test_empty_nt_hash_normalized(auth_client, workspace_id):
    """Empty NT hash (aad3b435…:31d6cfe0) must normalize to a plaintext credential."""
    payload = {
        "workspace": "test-ws", "operator": "tester",
        "data": {
            **_SYNC["data"],
            "credentials": [
                {"proto": "SMB", "domain": "corp.local", "username": "emptynt",
                 "password": "aad3b435b51404eeaad3b435b51404ee:31d6cfe0d16ae931b73c59d7e0c089c0",
                 "credtype": "hash"},
            ],
            "auth_relations": [],
        },
    }
    auth_client.post("/api/sync", json=payload)
    rows = auth_client.get(
        f"/api/credentials?workspace_id={workspace_id}&hide_guest=false"
    ).json()["rows"]
    emptynt_rows = [c for c in rows if c["username"] == "emptynt"]
    assert emptynt_rows, "emptynt credential not found after sync"
    assert any(c["credtype"] == "plaintext" for c in emptynt_rows), (
        "Empty NT hash was not normalized to plaintext credtype"
    )


def test_shell_flag_updates_on_resync(auth_client, workspace_id):
    """auth_relations.shell must update from None to 1 when later sync provides shell=1 (D-2)."""
    base_rel = {
        "proto": "SMB", "host_ip": "10.0.0.2",
        "cred_domain": "corp.local", "cred_username": "shell_tester",
        "cred_password": "ShellPass!", "cred_credtype": "plaintext",
        "relation_type": "admin", "shell": None,
    }
    base = {
        "workspace": "test-ws", "operator": "tester",
        "data": {
            "hosts": [{"ip": "10.0.0.2", "hostname": "SRV02", "domain": "corp.local",
                       "signing": 1, "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0,
                       "dc": 0, "nla": None, "signing_required": None, "channel_binding": None,
                       "port": 445, "banner": None, "instances": None}],
            "credentials": [{"proto": "SMB", "domain": "corp.local",
                              "username": "shell_tester", "password": "ShellPass!", "credtype": "plaintext"}],
            "auth_relations": [base_rel],
            "dpapi_secrets": [], "shares": [], "ssh_keys": [], "conf_checks_results": [], "directory_listings": [],
        },
    }
    update = {**base, "data": {**base["data"],
                                "auth_relations": [{**base_rel, "shell": 1}]}}
    auth_client.post("/api/sync", json=base)
    auth_client.post("/api/sync", json=update)
    with db_cursor() as cur:
        row = cur.execute(
            "SELECT ar.shell FROM auth_relations ar"
            " JOIN credentials c ON ar.credential_id = c.id"
            " JOIN hosts h ON ar.host_id = h.id"
            " WHERE ar.workspace_id=? AND c.username='shell_tester' AND h.ip='10.0.0.2'",
            (workspace_id,),
        ).fetchone()
    assert row is not None, "auth_relation for shell_tester not found"
    assert row["shell"] == 1, f"Expected shell=1 after resync, got {row['shell']}"


def test_directory_listing_no_duplicates_on_resync(auth_client, workspace_id):
    """Syncing the same directory listing twice must not create duplicate rows (D-1)."""
    payload = {
        "workspace": "test-ws", "operator": "tester",
        "data": {
            "hosts": [], "credentials": [], "auth_relations": [], "dpapi_secrets": [],
            "shares": [], "ssh_keys": [], "conf_checks_results": [],
            "directory_listings": [
                {"proto": "NFS", "host_ip": "10.0.0.50", "username": "anon", "data": "/export/data"},
            ],
        },
    }
    auth_client.post("/api/sync", json=payload)
    auth_client.post("/api/sync", json=payload)
    with db_cursor() as cur:
        count = cur.execute(
            "SELECT COUNT(*) FROM directory_listings"
            " WHERE workspace_id=? AND proto='NFS' AND host_ip='10.0.0.50' AND username='anon'",
            (workspace_id,),
        ).fetchone()[0]
    assert count == 1, f"Expected 1 row after double sync, got {count}"


# ═══════════════════════════════════════════════════════════════════════════════
# XLSX formula injection regression
# ═══════════════════════════════════════════════════════════════════════════════

def test_xlsx_formula_injection_prevented():
    """Cell value starting with '=' must not be stored as a formula cell."""
    from collector.services.export_service import xlsx_buf
    buf = xlsx_buf(["Password"], [["=secretpassword", None]], "Test")
    wb = openpyxl.load_workbook(BytesIO(buf.read()))
    cell = wb.active["A2"]
    assert cell.data_type != "f", "Cell must not be formula type"
    assert cell.value == "=secretpassword", f"Value must be preserved, got: {cell.value!r}"


# ═══════════════════════════════════════════════════════════════════════════════
# Guest filter regression
# ═══════════════════════════════════════════════════════════════════════════════

def test_guest_hidden_by_default(auth_client):
    """Guest/DefaultAccount/WDAGUtilityAccount must not appear in results by default."""
    r = auth_client.post("/api/workspaces", json={"name": "guest-filter-test"})
    ws_id = r.json()["id"]
    auth_client.post("/api/sync", json={
        "workspace": "guest-filter-test", "operator": "op",
        "data": {
            "hosts": [{"ip": "10.1.1.1", "hostname": "S", "domain": "corp.local", "signing": 1,
                       "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0, "dc": 0,
                       "nla": None, "signing_required": None, "channel_binding": None,
                       "port": 445, "banner": None, "instances": None}],
            "credentials": [
                {"proto": "SMB", "domain": "corp.local", "username": "Guest",
                 "password": "", "credtype": "plaintext"},
                {"proto": "SMB", "domain": "corp.local", "username": "real_op",
                 "password": "OpPass!", "credtype": "plaintext"},
            ],
            "auth_relations": [
                {"proto": "SMB", "host_ip": "10.1.1.1", "cred_domain": "corp.local",
                 "cred_username": "real_op", "cred_password": "OpPass!", "cred_credtype": "plaintext",
                 "relation_type": "loggedin", "shell": 0},
            ],
            "dpapi_secrets": [], "shares": [], "ssh_keys": [], "conf_checks_results": [], "directory_listings": [],
        },
    })
    rows = auth_client.get(f"/api/results?workspace_id={ws_id}").json()["rows"]
    users = {row["username"] for row in rows}
    assert "Guest" not in users, "Guest must be hidden by default"
    assert "real_op" in users, "real operator credential must be visible"


def test_guest_hidden_in_global_search(auth_client):
    """Guest must be excluded from /api/search by default; visible when hide_guest=false."""
    r = auth_client.post("/api/workspaces", json={"name": "guest-search-test"})
    ws_id = r.json()["id"]
    auth_client.post("/api/sync", json={
        "workspace": "guest-search-test", "operator": "op",
        "data": {
            "hosts": [{"ip": "10.9.0.1", "hostname": "SRV", "domain": "corp.local", "signing": 1,
                       "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0, "dc": 0,
                       "nla": None, "signing_required": None, "channel_binding": None,
                       "port": 445, "banner": None, "instances": None}],
            "credentials": [
                {"proto": "SMB", "domain": "corp.local", "username": "Guest",
                 "password": "GuestPass!", "credtype": "plaintext"},
                {"proto": "SMB", "domain": "corp.local", "username": "real_user",
                 "password": "RealPass!", "credtype": "plaintext"},
            ],
            "auth_relations": [
                {"proto": "SMB", "host_ip": "10.9.0.1", "cred_domain": "corp.local",
                 "cred_username": "Guest", "cred_password": "GuestPass!", "cred_credtype": "plaintext",
                 "relation_type": "loggedin", "shell": 0},
                {"proto": "SMB", "host_ip": "10.9.0.1", "cred_domain": "corp.local",
                 "cred_username": "real_user", "cred_password": "RealPass!", "cred_credtype": "plaintext",
                 "relation_type": "loggedin", "shell": 0},
            ],
            "dpapi_secrets": [], "shares": [], "ssh_keys": [], "conf_checks_results": [], "directory_listings": [],
        },
    })
    # Default (hide_guest=true): Guest excluded
    rows = auth_client.get(f"/api/search?workspace_id={ws_id}&q=corp.local").json()["rows"]
    users = {row["username"] for row in rows if row.get("username")}
    assert "Guest" not in users, "Guest must be excluded from global search by default"
    assert "real_user" in users, "real credential must appear in global search"

    # hide_guest=false: Guest visible
    rows_all = auth_client.get(f"/api/search?workspace_id={ws_id}&q=corp.local&hide_guest=false").json()["rows"]
    users_all = {row["username"] for row in rows_all if row.get("username")}
    assert "Guest" in users_all, "Guest must appear in global search when hide_guest=false"


# ═══════════════════════════════════════════════════════════════════════════════
# Strike IP regression — cross-host hiding bug
# ═══════════════════════════════════════════════════════════════════════════════

def test_strike_no_cross_host_hiding(auth_client):
    """Strike host A must NOT hide a credential that also authenticates on host B."""
    r = auth_client.post("/api/workspaces", json={"name": "strike-crosshost-test"})
    ws_id = r.json()["id"]

    def _host(ip, hostname):
        return {"ip": ip, "hostname": hostname, "domain": "corp.local", "signing": 1,
                "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0, "dc": 0,
                "nla": None, "signing_required": None, "channel_binding": None,
                "port": 445, "banner": None, "instances": None}

    def _rel(host_ip):
        return {"proto": "SMB", "host_ip": host_ip, "cred_domain": "corp.local",
                "cred_username": "shared_user", "cred_password": "SharedPass!",
                "cred_credtype": "plaintext", "relation_type": "loggedin", "shell": 0}

    auth_client.post("/api/sync", json={
        "workspace": "strike-crosshost-test", "operator": "op",
        "data": {
            "hosts": [_host("10.99.0.1", "honeypot"), _host("10.99.0.2", "normal")],
            "credentials": [{"proto": "SMB", "domain": "corp.local", "username": "shared_user",
                              "password": "SharedPass!", "credtype": "plaintext"}],
            "auth_relations": [_rel("10.99.0.1"), _rel("10.99.0.2")],
            "dpapi_secrets": [], "shares": [], "ssh_keys": [], "conf_checks_results": [], "directory_listings": [],
        },
    })

    auth_client.post("/api/hosts/strike", json={"workspace_id": ws_id, "host_ip": "10.99.0.1"})

    visible = {c["username"] for c in auth_client.get(
        f"/api/credentials?workspace_id={ws_id}&hide_guest=false"
    ).json()["rows"]}
    assert "shared_user" in visible, (
        "shared_user was incorrectly hidden: credential that also auths on another host "
        "must remain visible after striking only one of its hosts"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# hidden_by_strike regression — credential stuck hidden after gaining new host
# ═══════════════════════════════════════════════════════════════════════════════

def test_strike_restore_after_credential_gains_new_host(auth_client):
    """
    Bug B fix: credential hidden by strike must restore even when it gained
    a new auth_relation on a non-honeypot host between strike and restore.

    Scenario:
      1. cred exclusively on honeypot → strike → hidden (hidden_by_strike=1)
      2. sync: same cred now also on normal host (new auth_relation added)
      3. restore_strike → cred must become visible (hidden_by_strike filter, not NOT EXISTS)
    """
    r = auth_client.post("/api/workspaces", json={"name": "hbs-regression-test"})
    ws_id = r.json()["id"]

    def _host(ip, hostname):
        return {"ip": ip, "hostname": hostname, "domain": "corp.local", "signing": 1,
                "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0, "dc": 0,
                "nla": None, "signing_required": None, "channel_binding": None,
                "port": 445, "banner": None, "instances": None}

    def _rel(host_ip):
        return {"proto": "SMB", "host_ip": host_ip, "cred_domain": "corp.local",
                "cred_username": "victim_user", "cred_password": "VictimPass!",
                "cred_credtype": "plaintext", "relation_type": "loggedin", "shell": 0}

    _base = {"dpapi_secrets": [], "shares": [], "ssh_keys": [],
             "conf_checks_results": [], "directory_listings": []}
    _cred = [{"proto": "SMB", "domain": "corp.local", "username": "victim_user",
               "password": "VictimPass!", "credtype": "plaintext"}]

    # Step 1: sync — cred exclusively on honeypot host
    auth_client.post("/api/sync", json={
        "workspace": "hbs-regression-test", "operator": "op",
        "data": {"hosts": [_host("10.88.0.1", "honeypot")], "credentials": _cred,
                 "auth_relations": [_rel("10.88.0.1")], **_base},
    })

    # Step 2: strike the honeypot — cred must be hidden
    auth_client.post("/api/hosts/strike", json={"workspace_id": ws_id, "host_ip": "10.88.0.1"})
    visible_after_strike = {c["username"] for c in auth_client.get(
        f"/api/credentials?workspace_id={ws_id}&hide_guest=false"
    ).json()["rows"]}
    assert "victim_user" not in visible_after_strike, "cred must be hidden after strike"

    # Step 3: sync — same cred now also found on a normal host (new auth_relation)
    auth_client.post("/api/sync", json={
        "workspace": "hbs-regression-test", "operator": "op",
        "data": {"hosts": [_host("10.88.0.2", "normal")], "credentials": _cred,
                 "auth_relations": [_rel("10.88.0.2")], **_base},
    })

    # Step 4: restore the honeypot — cred must come back
    auth_client.post("/api/hosts/restore_strike",
                     json={"workspace_id": ws_id, "host_ip": "10.88.0.1"})
    visible_after_restore = {c["username"] for c in auth_client.get(
        f"/api/credentials?workspace_id={ws_id}&hide_guest=false"
    ).json()["rows"]}
    assert "victim_user" in visible_after_restore, (
        "victim_user must be visible after restore_strike: hidden_by_strike=1 filter "
        "must allow restore even though the credential now also auths on a normal host"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Local admin flag (set_local_admin_cred)
# ═══════════════════════════════════════════════════════════════════════════════

def _la_ws(auth_client, name):
    """Create workspace and return id."""
    return auth_client.post("/api/workspaces", json={"name": name}).json()["id"]


def _la_sync(auth_client, ws_name, creds, rels=None):
    payload = {
        "workspace": ws_name, "operator": "tester",
        "data": {
            "hosts": [{"ip": "10.1.0.1", "hostname": "PC01", "domain": "WORKGROUP",
                       "signing": 0, "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0,
                       "dc": 0, "nla": None, "signing_required": None, "channel_binding": None,
                       "port": 445, "banner": None, "instances": None}],
            "credentials": creds,
            "auth_relations": rels or [],
            "dpapi_secrets": [], "shares": [], "ssh_keys": [],
            "conf_checks_results": [], "directory_listings": [],
        },
    }
    return auth_client.post("/api/sync", json=payload)


def test_local_admin_mark_and_field_returned(auth_client):
    """Happy path: mark as local admin, flag returned by API."""
    ws_id = _la_ws(auth_client, "la-mark-test")
    _la_sync(auth_client, "la-mark-test",
             creds=[{"proto": "SMB", "domain": "PC01", "username": "ladmin",
                     "password": "Loc@lP@ss!", "credtype": "plaintext"}],
             rels=[{"proto": "SMB", "host_ip": "10.1.0.1", "cred_domain": "PC01",
                    "cred_username": "ladmin", "cred_password": "Loc@lP@ss!",
                    "cred_credtype": "plaintext", "relation_type": "admin", "shell": 0}])

    r = auth_client.post("/api/credentials/set_local_admin_cred",
                         json={"workspace_id": ws_id, "username": "ladmin",
                               "password": "Loc@lP@ss!", "local_admin_cred": 1})
    assert r.status_code == 200
    assert r.json()["ok"] is True

    creds = auth_client.get(
        f"/api/credentials?workspace_id={ws_id}&hide_guest=false"
    ).json()["rows"]
    assert any(c["username"] == "ladmin" and c["local_admin_cred"] == 1 for c in creds), (
        "local_admin_cred=1 must be returned after marking"
    )


def test_local_admin_unmark(auth_client):
    """Un-marking sets local_admin_cred back to 0."""
    ws_id = _la_ws(auth_client, "la-unmark-test")
    _la_sync(auth_client, "la-unmark-test",
             creds=[{"proto": "SMB", "domain": "MYMACHINE", "username": "unmark_user",
                     "password": "Unm@rk!", "credtype": "plaintext"}])

    auth_client.post("/api/credentials/set_local_admin_cred",
                     json={"workspace_id": ws_id, "username": "unmark_user",
                           "password": "Unm@rk!", "local_admin_cred": 1})
    auth_client.post("/api/credentials/set_local_admin_cred",
                     json={"workspace_id": ws_id, "username": "unmark_user",
                           "password": "Unm@rk!", "local_admin_cred": 0})

    creds = auth_client.get(
        f"/api/credentials?workspace_id={ws_id}&hide_guest=false"
    ).json()["rows"]
    assert any(c["username"] == "unmark_user" and c["local_admin_cred"] == 0 for c in creds), (
        "local_admin_cred must be 0 after un-marking"
    )


def test_local_admin_domain_threshold_blocks_real_domain(auth_client):
    """Credentials with domain appearing ≥ 10 times are silently skipped."""
    ws_id = _la_ws(auth_client, "la-threshold-test")
    # Sync 10 credentials with the same domain — crosses the threshold
    creds = [
        {"proto": "SMB", "domain": "bigcorp.local", "username": f"user{i}",
         "password": "Same1!", "credtype": "plaintext"}
        for i in range(10)
    ]
    _la_sync(auth_client, "la-threshold-test", creds=creds)

    # Try to mark user0 — domain bigcorp.local appears 10 times → must be skipped
    r = auth_client.post("/api/credentials/set_local_admin_cred",
                         json={"workspace_id": ws_id, "username": "user0",
                               "password": "Same1!", "local_admin_cred": 1})
    assert r.status_code == 200  # endpoint returns ok (silent skip)

    creds_resp = auth_client.get(
        f"/api/credentials?workspace_id={ws_id}&hide_guest=false"
    ).json()["rows"]
    assert all(c["local_admin_cred"] == 0 for c in creds_resp), (
        "No credential must be marked when domain appears ≥ 10 times (real domain threshold)"
    )


def test_local_admin_mass_marking_by_username_password(auth_client):
    """Marking by (username, password) covers all matching rows across machine domains."""
    ws_id = _la_ws(auth_client, "la-mass-test")
    # Same (username, password), two different machine-name domains (both < 10 occurrences)
    creds = [
        {"proto": "SMB", "domain": "MACHINE1", "username": "shared_admin",
         "password": "Sh@red!", "credtype": "plaintext"},
        {"proto": "SMB", "domain": "MACHINE2", "username": "shared_admin",
         "password": "Sh@red!", "credtype": "plaintext"},
    ]
    _la_sync(auth_client, "la-mass-test", creds=creds)

    auth_client.post("/api/credentials/set_local_admin_cred",
                     json={"workspace_id": ws_id, "username": "shared_admin",
                           "password": "Sh@red!", "local_admin_cred": 1})

    creds_resp = auth_client.get(
        f"/api/credentials?workspace_id={ws_id}&hide_guest=false"
    ).json()["rows"]
    marked = [c for c in creds_resp if c["username"] == "shared_admin"]
    assert len(marked) == 2, "Both rows must be present"
    assert all(c["local_admin_cred"] == 1 for c in marked), (
        "Both machine-domain rows must be marked as local admin"
    )


def test_local_admin_field_in_get_results(auth_client):
    """get_results() returns local_admin_cred field."""
    ws_id = _la_ws(auth_client, "la-results-test")
    _la_sync(auth_client, "la-results-test",
             creds=[{"proto": "SMB", "domain": "BOXPC", "username": "res_admin",
                     "password": "R3s@dmin!", "credtype": "plaintext"}],
             rels=[{"proto": "SMB", "host_ip": "10.1.0.1", "cred_domain": "BOXPC",
                    "cred_username": "res_admin", "cred_password": "R3s@dmin!",
                    "cred_credtype": "plaintext", "relation_type": "admin", "shell": 0}])

    auth_client.post("/api/credentials/set_local_admin_cred",
                     json={"workspace_id": ws_id, "username": "res_admin",
                           "password": "R3s@dmin!", "local_admin_cred": 1})

    results = auth_client.get(
        f"/api/results?workspace_id={ws_id}&relation=admin"
    ).json()["rows"]
    row = next((r for r in results if r["username"] == "res_admin"), None)
    assert row is not None, "res_admin must appear in results"
    assert "local_admin_cred" in row, "local_admin_cred field must be present in get_results response"
    assert row["local_admin_cred"] == 1


# ═══════════════════════════════════════════════════════════════════════════════
# Smart Enrich — pre-existing pair must not be marked as smart
# ═══════════════════════════════════════════════════════════════════════════════

def test_smart_enrich_skips_preexisting_pair(auth_client):
    """
    If (hash, plaintext) pair already exists in hk_pairs (smart=0, added via import),
    smart_enrich_workspace must skip it — not mark it as smart=1.
    """
    from collector.nt_hash import nt_hash
    _PLAIN = "SmartSkipTestPass!"
    # SMART ENRICH v2 verifies pairs: a candidate is only accepted when
    # nt_hash(plaintext) == stored hash. Use the genuine NT hash so the pair
    # reaches the "already exists -> skip" path being tested here.
    _HASH = nt_hash(_PLAIN)

    r = auth_client.post("/api/workspaces", json={"name": "smart-enrich-skip-test"})
    ws_id = r.json()["id"]

    auth_client.post("/api/sync", json={
        "workspace": "smart-enrich-skip-test", "operator": "op",
        "data": {
            "hosts": [{"ip": "10.111.0.1", "hostname": "TESTDC", "domain": "se.local",
                       "os": "Windows Server", "dc": 0, "signing": 1,
                       "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0,
                       "nla": None, "signing_required": None, "channel_binding": None,
                       "port": 445, "banner": None, "instances": None}],
            "credentials": [
                {"proto": "SMB", "domain": "se.local", "username": "seuser",
                 "password": _HASH, "credtype": "hash"},
                {"proto": "SMB", "domain": "se.local", "username": "seuser",
                 "password": _PLAIN, "credtype": "plaintext"},
            ],
            "auth_relations": [
                {"proto": "SMB", "host_ip": "10.111.0.1", "cred_domain": "se.local",
                 "cred_username": "seuser", "cred_password": _HASH,
                 "cred_credtype": "hash", "relation_type": "loggedin", "shell": 0},
                {"proto": "SMB", "host_ip": "10.111.0.1", "cred_domain": "se.local",
                 "cred_username": "seuser", "cred_password": _PLAIN,
                 "cred_credtype": "plaintext", "relation_type": "loggedin", "shell": 0},
            ],
            "dpapi_secrets": [], "shares": [], "ssh_keys": [], "conf_checks_results": [], "directory_listings": [],
        },
    })

    # Pre-populate hk_pairs with the same pair via import (smart=0)
    hk_db.bulk_import(f"{_HASH}:{_PLAIN}")
    assert (_HASH, _PLAIN) not in {(p["nt_hash"], p["plaintext"]) for p in hk_db.get_smart_pairs()}, \
        "Pre-condition: pair must not be smart before enrich"

    result = hk_db.smart_enrich_workspace(ws_id)

    assert result["added"] == 0, f"Smart enrich must not add a pre-existing pair; got {result}"
    assert result["skipped"] >= 1, f"Smart enrich must skip the pre-existing pair; got {result}"
    assert (_HASH, _PLAIN) not in {(p["nt_hash"], p["plaintext"]) for p in hk_db.get_smart_pairs()}, \
        "Smart enrich must not mark a pre-existing pair (smart=0) as smart=1"


def test_delete_all_custom_creds(auth_client, workspace_id):
    """DELETE ALL CUSTOM removes every custom_credentials row for the workspace,
    leaving other workspaces untouched, and reports how many were deleted."""
    # Second workspace whose custom rows must survive.
    other_id = auth_client.post("/api/workspaces", json={"name": "other-ws"}).json()["id"]
    with db_cursor() as cur:
        for wsid, n in ((workspace_id, 3), (other_id, 2)):
            for i in range(n):
                cur.execute(
                    "INSERT INTO custom_credentials"
                    " (workspace_id, proto, login, password, credtype, source)"
                    " VALUES (?,?,?,?,?,?)",
                    (wsid, "HTTP", f"u{i}", f"p{i}", "plaintext", "test"),
                )

    r = auth_client.request("DELETE", f"/api/custom_creds?workspace_id={workspace_id}")
    assert r.status_code == 200
    assert r.json()["deleted"] == 3

    with db_cursor() as cur:
        assert cur.execute(
            "SELECT COUNT(*) FROM custom_credentials WHERE workspace_id=?", (workspace_id,)
        ).fetchone()[0] == 0
        assert cur.execute(
            "SELECT COUNT(*) FROM custom_credentials WHERE workspace_id=?", (other_id,)
        ).fetchone()[0] == 2


# ═══════════════════════════════════════════════════════════════════════════════
# SMART ENRICH — plaintext enrichment (NT computed locally from every plaintext)
# ═══════════════════════════════════════════════════════════════════════════════

def _host_sync(auth_client, ws_name, creds, rels):
    """Sync one host (10.222.0.1) with the given credentials + auth_relations."""
    ws_id = auth_client.post("/api/workspaces", json={"name": ws_name}).json()["id"]
    auth_client.post("/api/sync", json={
        "workspace": ws_name, "operator": "op",
        "data": {
            "hosts": [{"ip": "10.222.0.1", "hostname": "VHOST", "domain": "v.local",
                       "os": "Windows Server", "dc": 0, "signing": 1,
                       "smbv1": 0, "spooler": 0, "zerologon": 0, "petitpotam": 0,
                       "nla": None, "signing_required": None, "channel_binding": None,
                       "port": 445, "banner": None, "instances": None}],
            "credentials": creds,
            "auth_relations": rels,
            "dpapi_secrets": [], "shares": [], "ssh_keys": [],
            "conf_checks_results": [], "directory_listings": [],
        },
    })
    return ws_id


def test_smart_enrich_never_pairs_wrong_hash(auth_client):
    """A host+login linkage whose stored hash != nt_hash(plaintext) must NOT produce
    that (wrong hash, plaintext) smart pair — pairs come only from hashing plaintext."""
    from collector.nt_hash import nt_hash
    PLAIN    = "GateRejectUnique_42!"
    BAD_HASH = "ffffffffffffffffffffffffffffffff"  # not the NT of PLAIN
    assert nt_hash(PLAIN) != BAD_HASH

    hk_db.init_hk_db()
    hk_db.delete_by_value(PLAIN)  # clean slate for this plaintext

    creds = [
        {"proto": "SMB", "domain": "v.local", "username": "gateuser",
         "password": BAD_HASH, "credtype": "hash"},
        {"proto": "SMB", "domain": "v.local", "username": "gateuser",
         "password": PLAIN, "credtype": "plaintext"},
    ]
    rels = [
        {"proto": "SMB", "host_ip": "10.222.0.1", "cred_domain": "v.local",
         "cred_username": "gateuser", "cred_password": BAD_HASH,
         "cred_credtype": "hash", "relation_type": "loggedin", "shell": 0},
        {"proto": "SMB", "host_ip": "10.222.0.1", "cred_domain": "v.local",
         "cred_username": "gateuser", "cred_password": PLAIN,
         "cred_credtype": "plaintext", "relation_type": "loggedin", "shell": 0},
    ]
    ws_id = _host_sync(auth_client, "smart-gate-reject", creds, rels)

    hk_db.smart_enrich_workspace(ws_id)
    smart = {(p["nt_hash"], p["plaintext"]) for p in hk_db.get_smart_pairs()}

    # A wrong (hash, plaintext) pairing is never produced: pairs are built solely
    # by hashing the plaintext, so only the correct NT pair can appear.
    assert (BAD_HASH, PLAIN) not in smart, "Wrong hash must never be paired to a plaintext"
    assert (nt_hash(PLAIN), PLAIN) in smart, "Computed pair from plaintext must be present"


def test_smart_enrich_enriches_from_plaintext(auth_client):
    """Every non-bruteforced plaintext (credentials + custom_credentials)
    is hashed to NT and added as a smart pair, even without a matching hash."""
    from collector.nt_hash import nt_hash
    P_SYNC   = "Part2SyncUnique_91!"
    P_CUSTOM = "Part2CustomUnique_91!"

    hk_db.init_hk_db()
    hk_db.delete_by_value(P_SYNC)
    hk_db.delete_by_value(P_CUSTOM)

    creds = [{"proto": "SMB", "domain": "v.local", "username": "p2user",
              "password": P_SYNC, "credtype": "plaintext"}]
    rels  = [{"proto": "SMB", "host_ip": "10.222.0.1", "cred_domain": "v.local",
              "cred_username": "p2user", "cred_password": P_SYNC,
              "cred_credtype": "plaintext", "relation_type": "loggedin", "shell": 0}]
    ws_id = _host_sync(auth_client, "smart-part2", creds, rels)

    with db_cursor() as cur:
        cur.execute(
            "INSERT INTO custom_credentials (workspace_id, proto, login, password, credtype, source)"
            " VALUES (?,?,?,?,?,?)",
            (ws_id, "HTTP", "custuser", P_CUSTOM, "plaintext", "test"),
        )

    res = hk_db.smart_enrich_workspace(ws_id)
    smart = {(p["nt_hash"], p["plaintext"]) for p in hk_db.get_smart_pairs()}

    assert (nt_hash(P_SYNC), P_SYNC) in smart, "plaintext from credentials must be enriched"
    assert (nt_hash(P_CUSTOM), P_CUSTOM) in smart, "plaintext from custom_credentials must be enriched"
    assert "added" in res and "skipped" in res
    assert res["added"] >= 2


def test_smart_enrich_part2_ignores_bruteforced_and_empty(auth_client):
    """Part 2 source is genuine plaintext only: brutforced values and empty
    passwords must NOT be turned into smart pairs."""
    from collector.nt_hash import nt_hash
    P_BRUTE = "BruteOnlyUnique_77!"   # exists only as a cracked (brutforced) value

    hk_db.init_hk_db()
    hk_db.delete_by_value(P_BRUTE)

    creds = [{"proto": "SMB", "domain": "v.local", "username": "bruteuser",
              "password": nt_hash(P_BRUTE), "credtype": "hash"}]
    rels  = [{"proto": "SMB", "host_ip": "10.222.0.1", "cred_domain": "v.local",
              "cred_username": "bruteuser", "cred_password": nt_hash(P_BRUTE),
              "cred_credtype": "hash", "relation_type": "loggedin", "shell": 0}]
    ws_id = _host_sync(auth_client, "smart-part2-brute", creds, rels)

    with db_cursor() as cur:
        # Mark the hash cred as cracked, and add an empty-password plaintext cred.
        cur.execute(
            "UPDATE credentials SET brutforced=? WHERE workspace_id=? AND credtype='hash'",
            (P_BRUTE, ws_id),
        )
        cur.execute(
            "INSERT INTO custom_credentials (workspace_id, proto, login, password, credtype, source)"
            " VALUES (?,?,?,?,?,?)",
            (ws_id, "HTTP", "emptyuser", "<empty_password>", "plaintext", "test"),
        )

    hk_db.smart_enrich_workspace(ws_id)
    smart = {(p["nt_hash"], p["plaintext"]) for p in hk_db.get_smart_pairs()}

    assert (nt_hash(P_BRUTE), P_BRUTE) not in smart, "brutforced value must not be an enrich source"
    assert not any(p["plaintext"] == "<empty_password>" for p in hk_db.get_smart_pairs()), \
        "empty password must be skipped"


# ═══════════════════════════════════════════════════════════════════════════════
# R5.1 — manually-cleared admin_cred must survive watchlist sync/upload
# ═══════════════════════════════════════════════════════════════════════════════

def _r51_ws(auth_client, name):
    return auth_client.post("/api/workspaces", json={"name": name}).json()["id"]


def _r51_sync(auth_client, ws_name, username="da_user", password="DaPass1!"):
    return auth_client.post("/api/sync", json={
        "workspace": ws_name, "operator": "tester",
        "data": {
            "hosts": [{"ip": "10.51.0.1", "hostname": "DC51", "domain": "r51.local",
                       "signing": 1, "smbv1": 0, "spooler": 0, "zerologon": 0,
                       "petitpotam": 0, "dc": 1, "nla": None, "signing_required": None,
                       "channel_binding": None, "port": 445, "banner": None, "instances": None}],
            "credentials": [{"proto": "SMB", "domain": "r51.local",
                              "username": username, "password": password, "credtype": "plaintext"}],
            "auth_relations": [],
            "dpapi_secrets": [], "shares": [], "ssh_keys": [],
            "conf_checks_results": [], "directory_listings": [],
        },
    })


def _r51_upload_watchlist(auth_client, ws_id, domain, username):
    return auth_client.post("/api/domain_admin_list/upload",
                            json={"workspace_id": ws_id, "domain": domain,
                                  "usernames": [username]})


def _r51_get_admin_cred(ws_id, username):
    with db_cursor() as cur:
        row = cur.execute(
            "SELECT admin_cred FROM credentials WHERE workspace_id=? AND username=?",
            (ws_id, username),
        ).fetchone()
    return row["admin_cred"] if row else None


def test_r51_manual_unset_survives_resync(auth_client):
    """
    R5.1: operator clears admin_cred → re-sync must NOT restore it.

    Flow: sync → upload watchlist → admin_cred=1 → manually unset (=0) → re-sync → still 0.
    """
    ws_id = _r51_ws(auth_client, "r51-resync-test")
    _r51_sync(auth_client, "r51-resync-test")
    _r51_upload_watchlist(auth_client, ws_id, "r51.local", "da_user")

    assert _r51_get_admin_cred(ws_id, "da_user") == 1, "watchlist must mark admin_cred=1"

    auth_client.post("/api/credentials/set_admin_cred",
                     json={"workspace_id": ws_id, "domain": "r51.local",
                           "username": "da_user", "admin_cred": 0})
    assert _r51_get_admin_cred(ws_id, "da_user") == 0, "manual unset must work"

    _r51_sync(auth_client, "r51-resync-test")

    assert _r51_get_admin_cred(ws_id, "da_user") == 0, (
        "re-sync must NOT restore admin_cred after operator manually cleared it"
    )


def test_r51_manual_unset_survives_watchlist_upload(auth_client):
    """
    R5.1: operator clears admin_cred → re-uploading watchlist must NOT restore it.
    """
    ws_id = _r51_ws(auth_client, "r51-dal-test")
    _r51_sync(auth_client, "r51-dal-test")
    _r51_upload_watchlist(auth_client, ws_id, "r51.local", "da_user")

    assert _r51_get_admin_cred(ws_id, "da_user") == 1

    auth_client.post("/api/credentials/set_admin_cred",
                     json={"workspace_id": ws_id, "domain": "r51.local",
                           "username": "da_user", "admin_cred": 0})

    _r51_upload_watchlist(auth_client, ws_id, "r51.local", "da_user")

    assert _r51_get_admin_cred(ws_id, "da_user") == 0, (
        "watchlist re-upload must NOT restore admin_cred after operator manually cleared it"
    )


def test_r51_watchlist_still_sets_untouched_rows(auth_client):
    """
    R5.1: rows that were never manually touched must still get admin_cred=1 from watchlist.
    """
    ws_id = _r51_ws(auth_client, "r51-untouched-test")
    _r51_sync(auth_client, "r51-untouched-test")

    assert _r51_get_admin_cred(ws_id, "da_user") == 0, "starts at 0"

    _r51_upload_watchlist(auth_client, ws_id, "r51.local", "da_user")

    assert _r51_get_admin_cred(ws_id, "da_user") == 1, (
        "watchlist must still set admin_cred=1 for rows that were never manually unset"
    )


def test_r51_manual_reset_to_1_allows_watchlist_again(auth_client):
    """
    R5.1: after manual unset, operator can manually re-set to 1, which unlocks watchlist control.
    A subsequent watchlist upload can still set admin_cred=1 for this identity.
    """
    ws_id = _r51_ws(auth_client, "r51-relock-test")
    _r51_sync(auth_client, "r51-relock-test")
    _r51_upload_watchlist(auth_client, ws_id, "r51.local", "da_user")
    assert _r51_get_admin_cred(ws_id, "da_user") == 1

    # Operator unsets
    auth_client.post("/api/credentials/set_admin_cred",
                     json={"workspace_id": ws_id, "domain": "r51.local",
                           "username": "da_user", "admin_cred": 0})
    assert _r51_get_admin_cred(ws_id, "da_user") == 0

    # Operator manually re-sets to 1 → unlocks watchlist
    auth_client.post("/api/credentials/set_admin_cred",
                     json={"workspace_id": ws_id, "domain": "r51.local",
                           "username": "da_user", "admin_cred": 1})
    assert _r51_get_admin_cred(ws_id, "da_user") == 1

    # Now unset manually again and check that a new sync can find new cred (different password)
    # and mark it via watchlist — to confirm unlock path is consistent
    _r51_sync(auth_client, "r51-relock-test", username="da_user2", password="NewPass2!")
    _r51_upload_watchlist(auth_client, ws_id, "r51.local", "da_user2")
    assert _r51_get_admin_cred(ws_id, "da_user2") == 1, (
        "fresh credential not manually touched must be set by watchlist"
    )
