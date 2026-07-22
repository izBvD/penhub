"""
Toolbox API routes: /api/toolbox/*

Block 1 — Custom import (XLSX template download; future: upload & apply to workspace).
Block 2 — NXCExtractor list exports (logins, passwords, hashes, IPs, spray archive).
Block 3 — Operator environment config (scripts download, config strings, BH workspace config).
"""

import io
import sqlite3
import zipfile
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

import collector.core.auth as _auth
from collector.core.auth import verify_token
from collector.db import db_cursor

router = APIRouter()


_SCRIPT_FILES = ["nxc_collector", "nxce.py", "nxc_updater.py", "dicgenerat.py", "collector_dc.py", "collector_hosts.py"]


# ── Block 1: Custom import ────────────────────────────────────────────────────

_CUSTOM_IMPORT_COLUMNS = [
    # (header_name, hint_text, required)
    ("Proto",    "SMB / RDP / SSH / etc.",                False),
    ("IP",       "target IP address",                     False),
    ("Port",     "integer, e.g. 445",                     False),
    ("Domain",   "AD domain or workgroup",                False),
    ("Login",    "username — Login OR Password required",  True),
    ("Password", "plaintext or hash — req. if no Login",  True),
    ("Type",     "hash / plaintext  (blank → plaintext)", False),
    ("URL",      "resource URL (web, vpn, etc.)",         False),
    ("Source",   "leak / phishing / manual…",             False),
    ("Comment",  "free-form comment",                     False),
]

# Header name (case-insensitive) → internal field key
_CUSTOM_IMPORT_COL_MAP = {
    "proto":    "proto",
    "ip":       "ip",
    "port":     "port",
    "domain":   "domain",
    "login":    "login",
    "password": "password",
    "type":     "credtype",
    "url":      "url",
    "source":   "source",
    "comment":  "comment",
}

_UPLOAD_XLSX_MAX_BYTES = 20 * 1024 * 1024   # 20 MB


def _parse_custom_import_xlsx(data: bytes) -> tuple:
    """
    Parse XLSX bytes → (rows, matched_headers, unrecognized_headers).
    Headers matched case-insensitively against _CUSTOM_IMPORT_COL_MAP.
    matched_headers: original-case names of recognized columns.
    unrecognized_headers: non-empty header names that had no match.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], [], []

    # Build col_index → field_key mapping from header row
    col_map: dict = {}
    matched_headers:      list = []
    unrecognized_headers: list = []
    for idx, cell in enumerate(rows[0]):
        name = str(cell or "").strip()
        if not name:
            continue
        if name.lower() in _CUSTOM_IMPORT_COL_MAP:
            col_map[idx] = _CUSTOM_IMPORT_COL_MAP[name.lower()]
            matched_headers.append(name)
        else:
            unrecognized_headers.append(name)

    result = []
    for row in rows[1:]:
        d: dict = {}
        for idx, field in col_map.items():
            d[field] = row[idx] if idx < len(row) else None
        result.append(d)
    return result, matched_headers, unrecognized_headers


def _import_custom_row(cur, workspace_id: int, row: dict) -> str:
    """
    Import one parsed row into custom_credentials.
    Returns: 'added' | 'enriched' | 'already_existed' | 'skipped'

    Key fields (part of UNIQUE index): proto, ip, port, domain, url.
    Meta fields (not in UNIQUE key): source, comment.
    Enrichment rules:
      - NULL/empty stored key field + non-empty incoming → UPDATE (enrich).
      - Non-NULL stored key field ≠ incoming value → incompatible candidate, skip (may INSERT).
      - Meta fields: same enrichment rule, no conflict check needed.
    """
    def _s(v) -> Optional[str]:
        if v is None:
            return None
        s = str(v).strip()
        return s or None

    def _i(v) -> Optional[int]:
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    login    = _s(row.get("login"))    or ""
    password = _s(row.get("password")) or ""

    if not login and not password:
        return "skipped"

    credtype = "hash" if (_s(row.get("credtype")) or "").lower() == "hash" else "plaintext"
    proto    = _s(row.get("proto"))
    ip       = _s(row.get("ip"))
    port     = _i(row.get("port"))
    domain   = _s(row.get("domain"))
    url      = _s(row.get("url"))
    source   = _s(row.get("source"))
    comment  = _s(row.get("comment"))

    # ── Step 1: exact UNIQUE key match ───────────────────────────────────────
    exact = cur.execute("""
        SELECT id, source, comment FROM custom_credentials
        WHERE workspace_id=?
          AND COALESCE(proto,'')  = COALESCE(?,'')
          AND COALESCE(ip,'')     = COALESCE(?,'')
          AND COALESCE(port,-1)   = COALESCE(?,-1)
          AND COALESCE(domain,'') = COALESCE(?,'')
          AND login=? AND password=? AND credtype=?
          AND COALESCE(url,'')    = COALESCE(?,'')
    """, (workspace_id, proto, ip, port, domain, login, password, credtype, url)).fetchone()

    if exact:
        # url is a key field — only meta fields can be enriched at this step
        updates: dict = {}
        if source  and not exact["source"]:  updates["source"]  = source
        if comment and not exact["comment"]: updates["comment"] = comment
        if updates:
            set_clause = ", ".join(f"{k}=?" for k in updates)
            cur.execute(
                f"UPDATE custom_credentials SET {set_clause},"
                " updated_at=strftime('%Y-%m-%dT%H:%M:%SZ','now'),"
                " imported_at=strftime('%Y-%m-%dT%H:%M:%SZ','now')"
                " WHERE id=?",
                (*updates.values(), exact["id"]),
            )
            return "enriched"
        return "already_existed"

    # ── Step 2: enrichable match (same login+password+credtype, some key fields NULL) ──
    candidates = cur.execute("""
        SELECT id, proto, ip, port, domain, url, source, comment
        FROM custom_credentials
        WHERE workspace_id=? AND login=? AND password=? AND credtype=?
        ORDER BY id
    """, (workspace_id, login, password, credtype)).fetchall()

    for c in candidates:
        # Compatibility check: all already-set key fields must match (or incoming is empty).
        # If any set field differs → these are separate credentials → skip this candidate.
        if (c["proto"]            and proto            and c["proto"]  != proto):  continue
        if (c["ip"]               and ip               and c["ip"]     != ip):     continue
        if (c["port"] is not None and port is not None and c["port"]   != port):   continue
        if (c["domain"]           and domain           and c["domain"] != domain): continue
        if (c["url"]              and url              and c["url"]    != url):     continue

        # Compute merged key fields: fill NULL/empty stored values with incoming ones
        new_proto  = proto  if (proto  and not c["proto"])              else c["proto"]
        new_ip     = ip     if (ip     and not c["ip"])                 else c["ip"]
        new_port   = port   if (port is not None and c["port"] is None) else c["port"]
        new_domain = domain if (domain and not c["domain"])             else c["domain"]
        new_url    = url    if (url    and not c["url"])                 else c["url"]
        # Compute merged meta fields
        new_source  = source  if (source  and not c["source"])  else c["source"]
        new_comment = comment if (comment and not c["comment"]) else c["comment"]

        key_changed  = (new_proto != c["proto"] or new_ip != c["ip"] or
                        new_port != c["port"] or new_domain != c["domain"] or
                        new_url != c["url"])
        meta_changed = (new_source != c["source"] or new_comment != c["comment"])

        if not key_changed and not meta_changed:
            continue   # truly identical — will be caught by step 1 on re-import

        if key_changed:
            # Verify the merged key doesn't conflict with another existing row
            conflict = cur.execute("""
                SELECT id FROM custom_credentials
                WHERE workspace_id=?
                  AND COALESCE(proto,'')  = COALESCE(?,'')
                  AND COALESCE(ip,'')     = COALESCE(?,'')
                  AND COALESCE(port,-1)   = COALESCE(?,-1)
                  AND COALESCE(domain,'') = COALESCE(?,'')
                  AND login=? AND password=? AND credtype=?
                  AND COALESCE(url,'')    = COALESCE(?,'')
                  AND id != ?
            """, (workspace_id, new_proto, new_ip, new_port, new_domain,
                  login, password, credtype, new_url, c["id"])).fetchone()
            if conflict:
                continue   # try next candidate

        cur.execute("""
            UPDATE custom_credentials
            SET proto=?, ip=?, port=?, domain=?,
                url=?, source=?, comment=?,
                updated_at  = strftime('%Y-%m-%dT%H:%M:%SZ','now'),
                imported_at = strftime('%Y-%m-%dT%H:%M:%SZ','now')
            WHERE id=?
        """, (new_proto, new_ip, new_port, new_domain,
              new_url, new_source, new_comment, c["id"]))
        return "enriched"

    # ── Step 3: INSERT as new credential ────────────────────────────────────
    try:
        cur.execute("""
            INSERT INTO custom_credentials
                (workspace_id, proto, ip, port, domain,
                 login, password, credtype, url, source, comment)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (workspace_id, proto, ip, port, domain,
               login, password, credtype, url, source, comment))
        return "added"
    except sqlite3.IntegrityError:
        return "already_existed"


@router.get("/api/toolbox/custom-import/template", dependencies=[Depends(verify_token)])
def custom_import_template():
    """Download XLSX template for custom credential import."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL   = PatternFill("solid", fgColor="2F75B6")  # standard blue
    REQUIRED_FILL = PatternFill("solid", fgColor="C55A11")  # orange — required (Login | Password)
    HEADER_FONT   = Font(bold=True, color="FFFFFF")
    HINT_FILL     = PatternFill("solid", fgColor="F2F2F2")
    HINT_FONT     = Font(italic=True, color="595959")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Custom Import"

    for col_idx, (name, hint, required) in enumerate(_CUSTOM_IMPORT_COLUMNS, start=1):
        hdr = ws.cell(row=1, column=col_idx, value=name)
        hdr.fill   = REQUIRED_FILL if required else HEADER_FILL
        hdr.font   = HEADER_FONT
        hdr.alignment = Alignment(horizontal="center")

        tip = ws.cell(row=2, column=col_idx, value=hint)
        tip.fill   = HINT_FILL
        tip.font   = HINT_FONT
        tip.alignment = Alignment(horizontal="left")

        width = max(len(name), len(hint)) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=custom_import_template.xlsx"},
    )


@router.post("/api/toolbox/custom-import/upload", dependencies=[Depends(verify_token)])
async def custom_import_upload(
    workspace_id: int = Form(...),
    file: UploadFile = File(...),
):
    """Import credentials from a filled XLSX template into custom_credentials."""
    raw = await file.read(_UPLOAD_XLSX_MAX_BYTES + 1)
    if len(raw) > _UPLOAD_XLSX_MAX_BYTES:
        raise HTTPException(status_code=413, detail="File too large — max 20 MB")

    with db_cursor() as cur:
        if not cur.execute("SELECT id FROM workspaces WHERE id=?", (workspace_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Workspace not found")

    try:
        rows, matched_headers, unrecognized_headers = _parse_custom_import_xlsx(raw)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Cannot parse XLSX: {exc}")

    added = enriched = already_existed = skipped = 0
    with db_cursor() as cur:
        for row in rows:
            result = _import_custom_row(cur, workspace_id, row)
            if   result == "added":           added           += 1
            elif result == "enriched":        enriched        += 1
            elif result == "already_existed": already_existed += 1
            else:                             skipped         += 1

    return {
        "added":                added,
        "enriched":             enriched,
        "already_existed":      already_existed,
        "skipped":              skipped,
        "total_rows":           len(rows),
        "matched_headers":      matched_headers,
        "unrecognized_headers": unrecognized_headers,
    }


# ── Block 3: Operator environment config ─────────────────────────────────────

@router.get("/api/toolbox/scripts", dependencies=[Depends(verify_token)])
def toolbox_scripts():
    """Download ZIP with operator scripts: nxc_collector, nxce.py, nxc_updater.py."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in _SCRIPT_FILES:
            p = Path(name)
            if p.exists():
                zf.write(str(p), name)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=operator_scripts.zip"},
    )


@router.get("/api/toolbox/server-info", dependencies=[Depends(verify_token)])
def toolbox_server_info(request: Request, workspace_id: Optional[int] = Query(None)):
    """Return server URL, port, password, and workspace name for config string generation."""
    base     = request.base_url
    hostname = base.hostname or ""
    port     = base.port or (443 if base.scheme == "https" else 80)
    server_url = f"{base.scheme}://{hostname}"

    ws_name = ""
    if workspace_id is not None:
        with db_cursor() as cur:
            row = cur.execute(
                "SELECT name FROM workspaces WHERE id = ?", (workspace_id,)
            ).fetchone()
            if row:
                ws_name = row[0]

    return {
        "server_url": server_url,
        "port": port,
        "password": _auth.APP_PASSWORD,
        "ws_name": ws_name,
    }


# ── Workspace config (per-workspace BH settings) ─────────────────────────────

class WsConfigBody(BaseModel):
    bh_ip:     Optional[str] = None
    bh_login:  Optional[str] = None
    bh_pass:   Optional[str] = None
    bh_port:   Optional[str] = None
    bh_enable: Optional[str] = None


_BH_DEFAULTS = {
    "bh_login":  "neo4j",
    "bh_pass":   "bloodhoundcommunityedition",
    "bh_port":   "7687",
    "bh_enable": "true",
}


@router.get("/api/toolbox/ws-config/{ws_id}", dependencies=[Depends(verify_token)])
def toolbox_get_ws_config(ws_id: int):
    """Return per-workspace config, merged with BH defaults for unset keys."""
    with db_cursor() as cur:
        rows = cur.execute(
            "SELECT key, value FROM workspace_config WHERE workspace_id = ?", (ws_id,)
        ).fetchall()
    cfg = dict(_BH_DEFAULTS)
    cfg.update({r[0]: r[1] for r in rows})
    return cfg


@router.put("/api/toolbox/ws-config/{ws_id}", dependencies=[Depends(verify_token)])
def toolbox_put_ws_config(ws_id: int, body: WsConfigBody):
    """Save per-workspace config fields (upsert)."""
    data = {k: v for k, v in body.model_dump().items() if v is not None and v.strip() != ''}
    with db_cursor() as cur:
        for key, value in data.items():
            cur.execute("""
                INSERT INTO workspace_config(workspace_id, key, value) VALUES(?, ?, ?)
                ON CONFLICT(workspace_id, key) DO UPDATE SET value = excluded.value
            """, (ws_id, key, value))
    return {"ok": True}
