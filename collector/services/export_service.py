"""
XLSX generation helpers.
All functions are pure (no side effects, no DB access).
"""

import re
from io import BytesIO


# openpyxl rejects XML-illegal chars (e.g. NUL bytes in NXC OS strings)
_ILLEGAL_XML = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]')


def sanitize_cell(v):
    """Strip XML-illegal characters from string cell values."""
    if isinstance(v, str):
        return _ILLEGAL_XML.sub('', v)
    return v


def xlsx_buf(headers: list, rows: list, sheet_name: str = "Data") -> BytesIO:
    """
    Build an openpyxl workbook and return it as a BytesIO buffer.

    Each row may have an optional trailing tag string (e.g. "__admin__", "__hash__", "__dpapi__")
    used ONLY for row fill color — it is stripped before writing cell data.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="2F75B6")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ADMIN_FILL  = PatternFill("solid", fgColor="FFD7D7")
    HASH_FILL   = PatternFill("solid", fgColor="FFF2CC")
    DPAPI_FILL  = PatternFill("solid", fgColor="E2D9F3")
    CUSTOM_FILL = PatternFill("solid", fgColor="FFF3E0")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        # Last element is a row-type tag (string like "__admin__" or None) used only for
        # row fill color. It must be stripped before writing cell data.
        has_tag = bool(row) and isinstance(row[-1], (str, type(None))) and (
            row[-1] is None or (isinstance(row[-1], str) and row[-1].startswith("__"))
        )
        row_type = row[-1] if has_tag else ""
        data_row = row[:-1] if has_tag else row
        ws.append([sanitize_cell(v) for v in data_row])
        last_row = ws.max_row
        # openpyxl marks strings starting with '=' as formula cells (data_type='f').
        # Excel cannot evaluate these and removes them: "Removed Records: Formula".
        # Restore original string and force string type on any such cell.
        for cell in ws[last_row]:
            if cell.data_type == 'f':
                cell.data_type = 's'
        if row_type == "__admin__":
            fill = ADMIN_FILL
        elif row_type == "__hash__":
            fill = HASH_FILL
        elif row_type == "__dpapi__":
            fill = DPAPI_FILL
        elif row_type == "__custom__":
            fill = CUSTOM_FILL
        else:
            fill = None
        if fill is not None:
            for cell in ws[last_row]:
                cell.fill = fill

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def allcred_xlsx(plain_rows: list, hash_rows: list, dpapi_rows: list,
                 custom_rows: list | None = None,
                 local_admin_rows: list | None = None) -> BytesIO:
    """
    Sectioned ALL CRED workbook:
      1. Plaintext credentials  — sorted by domain
      2. LOCAL ADMIN credentials — per-machine, sorted by machine then login (separator + headers)
      3. Hash credentials       — sorted by domain  (separator + repeated headers)
      4. DPAPI credentials      — sorted by host    (separator + repeated headers)
      5. Custom credentials     — sorted by proto   (separator + repeated headers)

    local_admin_rows: deduped AFTER regular dedup so each machine keeps its own row.
    Domain column shows real machine name (e.g. "DESKTOP-ABC123"), not "local admin".
    The section separator "LOCAL ADMIN" provides the visual label.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL      = PatternFill("solid", fgColor="2F75B6")
    HEADER_FONT      = Font(bold=True, color="FFFFFF")
    HASH_FILL        = PatternFill("solid", fgColor="FFF2CC")
    DPAPI_FILL       = PatternFill("solid", fgColor="E2D9F3")
    CUSTOM_FILL      = PatternFill("solid", fgColor="FFF3E0")
    LOCAL_ADMIN_FILL = PatternFill("solid", fgColor="DCE6F1")  # light blue
    SEP_FILL         = PatternFill("solid", fgColor="404040")
    SEP_FONT         = Font(bold=True, color="FFFFFF")
    N_COLS           = 8  # widest section (custom) = 8 cols

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credentials"

    def _header(labels):
        ws.append(labels)
        for cell in ws[ws.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

    def _separator(title):
        ws.append([title] + [""] * (N_COLS - 1))
        row_idx = ws.max_row
        ws.merge_cells(f"A{row_idx}:{get_column_letter(N_COLS)}{row_idx}")
        for cell in ws[row_idx]:
            cell.fill = SEP_FILL
            cell.font = SEP_FONT
            cell.alignment = Alignment(horizontal="center")

    def _data_row(values, fill=None):
        ws.append([sanitize_cell(v) for v in values])
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            if cell.data_type == 'f':
                cell.data_type = 's'
            if fill:
                cell.fill = fill

    # ── Section 1: Plaintext ─────────────────────────────────────────────────
    _header(["Service", "Domain", "Login", "Password"])
    for r in plain_rows:
        _data_row([r.get("proto"), r.get("domain"), r.get("username"), r.get("password")])

    # ── Section 2: Local Admin ───────────────────────────────────────────────
    # Rows deduped with real machine names as domain — each machine appears separately.
    if local_admin_rows:
        _separator("LOCAL ADMIN")
        _header(["Service", "Machine", "Login", "Password", "Type"])
        for r in local_admin_rows:
            _data_row([r.get("proto"), r.get("domain"), r.get("username"),
                       r.get("password"), r.get("credtype")],
                      fill=LOCAL_ADMIN_FILL)

    # ── Section 3: Hashes ────────────────────────────────────────────────────
    if hash_rows:
        _separator("HASHES")
        _header(["Service", "Domain", "Login", "Hash"])
        for r in hash_rows:
            _data_row([r.get("proto"), r.get("domain"), r.get("username"), r.get("password")],
                      fill=HASH_FILL)

    # ── Section 3: DPAPI ─────────────────────────────────────────────────────
    if dpapi_rows:
        _separator("DPAPI")
        _header(["Service", "Host", "URL", "Login", "Password"])
        for r in dpapi_rows:
            _data_row([r.get("proto"), r.get("host_ip"), r.get("url"), r.get("username"), r.get("password")],
                      fill=DPAPI_FILL)

    # ── Section 4: Custom ────────────────────────────────────────────────────
    if custom_rows:
        _separator("CUSTOM")
        _header(["Proto", "IP", "Domain", "Login", "Password", "URL", "Source", "Comment"])
        for r in custom_rows:
            fill = CUSTOM_FILL if r.get("credtype") == "hash" else None
            _data_row([r.get("proto"), r.get("ip"), r.get("domain"),
                       r.get("username"), r.get("password"),
                       r.get("url"), r.get("source"), r.get("comment")],
                      fill=fill)

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def cred_xlsx(rows_data: list, headers: list | None = None) -> BytesIO:
    """
    Build the ALL CRED workbook (Service / Domain / Login / Password / Brutforced / Is Hash?).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="2F75B6")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HASH_FILL   = PatternFill("solid", fgColor="FFF2CC")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credentials"

    headers = headers or ["Service", "Domain", "Login", "Password", "Is Hash?"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r in rows_data:
        pw = r.get("brutforced") or r["password"]
        is_hash = "+" if r.get("credtype") == "hash" and not r.get("brutforced") else ""
        row = [r["proto"], r["domain"], r["username"], pw, is_hash]
        ws.append([sanitize_cell(v) for v in row])
        for cell in ws[ws.max_row]:
            if cell.data_type == 'f':
                cell.data_type = 's'
        if r.get("credtype") == "hash" and not r.get("brutforced"):
            for cell in ws[ws.max_row]:
                cell.fill = HASH_FILL

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
